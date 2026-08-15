from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from apps.engine.parsers.bank.money import normalize_date
from apps.engine.validators.gstin import gstin_flags

# Canonical field → Zoho Books export header aliases (normalized, first match wins).
HEADER_MAP: dict[str, tuple[str, ...]] = {
    "invoice_number": (
        "invoice number",
        "invoice_number",
        "invoice no",
        "invoice no.",
        "invoice #",
        "bill number",
        "bill no",
        "bill no.",
        "bill #",
    ),
    "invoice_date": (
        "invoice date",
        "invoice_date",
        "bill date",
        "date",
    ),
    "supplier_name": (
        "customer name",
        "vendor name",
        "display name",
        "supplier name",
    ),
    "supplier_gstin": (
        "gst identification number (gstin)",
        "gst identification number",
        "gstin",
        "gst no",
        "gst no.",
    ),
    "taxable_value": (
        "subtotal",
        "sub total",
        "taxable",
        "taxable value",
        "taxable amount",
    ),
    "tax_amount": (
        "tax amount",
        "item tax amount",
        "total tax amount",
        "tax",
    ),
    "tax_percent": (
        "item tax %",
        "item tax%",
        "tax %",
        "tax%",
        "tax percent",
        "item tax percent",
    ),
    "invoice_value": (
        "total",
        "invoice_value",
        "invoice value",
        "grand total",
    ),
    "hsn": (
        "hsn/sac",
        "hsn",
        "sac",
        "hsn/sac code",
        "hsn sac",
    ),
}

_SALES_HEADERS = {
    "invoice number",
    "invoice_number",
    "invoice no",
    "invoice no.",
    "invoice #",
}
_ISO_DATE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")
_SPREADSHEET = {".xlsx", ".xls"}


def parse_zoho_file(path: Path, filename: str | None = None) -> dict:
    try:
        return _parse(Path(path), filename)
    except Exception:
        return {"rows": []}


def _parse(path: Path, filename: str | None) -> dict:
    source = filename or path.name
    suffix = path.suffix.lower()
    if suffix in _SPREADSHEET:
        table = _read_spreadsheet(path)
    else:
        table = _read_csv(path)
    if not table:
        return {"rows": []}
    headers, records = table
    register = _detect_register(headers)
    rows = []
    for record in records:
        row = _canonical_row(record, register, source)
        if row is not None:
            rows.append(row)
    return {"rows": rows}


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]] | None:
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            if not sample.strip():
                return None
            handle.seek(0)
            reader = csv.DictReader(handle)
            headers = [str(name) for name in (reader.fieldnames or []) if name]
            if not headers:
                return None
            records = [dict(line) for line in reader]
    except (OSError, UnicodeError, csv.Error):
        return None
    return headers, records


def _read_spreadsheet(path: Path) -> tuple[list[str], list[dict[str, Any]]] | None:
    if path.suffix.lower() == ".xls":
        return _read_xls(path)
    return _read_xlsx(path)


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]] | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            return None
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    except Exception:
        return None
    finally:
        workbook.close()
    return _table_from_matrix(matrix)


def _read_xls(path: Path) -> tuple[list[str], list[dict[str, Any]]] | None:
    try:
        import xlrd
    except ImportError:
        return None
    try:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        matrix = [sheet.row_values(index) for index in range(sheet.nrows)]
    except Exception:
        return None
    return _table_from_matrix(matrix)


def _table_from_matrix(matrix: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]] | None:
    header_idx = None
    for index, row in enumerate(matrix):
        if _row_has_value(row):
            header_idx = index
            break
    if header_idx is None:
        return None
    headers = [_header_label(cell) for cell in matrix[header_idx]]
    if not any(headers):
        return None
    records: list[dict[str, Any]] = []
    for row in matrix[header_idx + 1 :]:
        if not _row_has_value(row):
            continue
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        records.append(record)
    return headers, records


def _detect_register(headers: list[str]) -> str | None:
    norms = [_norm(header) for header in headers if header]
    if any(name in _SALES_HEADERS for name in norms):
        return "sales"
    for name in norms:
        if "bill number" in name or name in {"bill no", "bill no.", "bill #"}:
            return "purchase"
        if "vendor" in name or "purchase" in name:
            return "purchase"
    return None


def _canonical_row(record: dict[str, Any], register: str | None, source: str) -> dict | None:
    lookup = {_norm(key): value for key, value in record.items() if key}
    invoice_number = _text(_pick(lookup, HEADER_MAP["invoice_number"]))
    invoice_date = _invoice_date(_pick(lookup, HEADER_MAP["invoice_date"]))
    supplier_name = _text(_pick(lookup, HEADER_MAP["supplier_name"]))
    gstin = _text(_pick(lookup, HEADER_MAP["supplier_gstin"]))
    taxable_value = _num(_pick(lookup, HEADER_MAP["taxable_value"]))
    invoice_value = _num(_pick(lookup, HEADER_MAP["invoice_value"]))
    hsn = _text(_pick(lookup, HEADER_MAP["hsn"]))
    tax = _tax(lookup, taxable_value)
    if not any((invoice_number, invoice_date, supplier_name, gstin, taxable_value, tax, invoice_value, hsn)):
        return None
    return {
        "register": register,
        "supplier_name": supplier_name,
        "supplier_gstin": gstin,
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "taxable_value": taxable_value,
        "tax": tax,
        "invoice_value": invoice_value,
        "hsn": hsn,
        "flags": gstin_flags(gstin) if gstin else [],
        "source": source,
    }


def _tax(lookup: dict[str, Any], taxable_value: float | None) -> float | None:
    amount = _num(_pick(lookup, HEADER_MAP["tax_amount"]))
    if amount is not None:
        return amount
    percent = _num(_pick(lookup, HEADER_MAP["tax_percent"]))
    if percent is not None and taxable_value is not None:
        return round(taxable_value * percent / 100.0, 2)
    return None


def _pick(lookup: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        if alias in lookup:
            value = lookup[alias]
            if value is not None and value != "":
                return value
    return None


def _invoice_date(raw: Any) -> str | None:
    if raw is None or raw == "":
        return None
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    text = str(raw).strip()
    if not text:
        return None
    iso = _ISO_DATE.match(text)
    if iso:
        return f"{iso.group(1)}-{iso.group(2)}-{iso.group(3)}"
    parts = re.split(r"[/\-.]", text)
    if len(parts) == 3 and len(parts[0]) == 4 and parts[0].isdigit():
        try:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
        except ValueError:
            return text
    return normalize_date(text)


def _num(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).replace("₹", "").replace(",", "").replace("%", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _text(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _norm(name: str) -> str:
    return " ".join(str(name).strip().lower().split())


def _header_label(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def _row_has_value(row: list[Any]) -> bool:
    for cell in row:
        if cell is None or cell == "":
            continue
        if str(cell).strip():
            return True
    return False
