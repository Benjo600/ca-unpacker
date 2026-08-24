from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from apps.engine.clients import create_client
from apps.engine.dump import ingest_paths, list_period_files, start_job
from apps.engine.firm import save_firm
from apps.engine.periods import create_period


class VisibleParseFailureTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()
        save_firm("Test firm")
        self.period = create_period(create_client("Acme")["id"], "Jul 2026")

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def test_gstr_parser_exception_is_on_file_and_job_warnings(self) -> None:
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        bad = inbox / "GSTR-2B_July.json"
        bad.write_text('{"gstin":"27AAPFU0939F1ZV"}', encoding="utf-8")
        job = start_job(self.period["id"])
        with patch("apps.engine.pipeline.parse_gstr_file", side_effect=RuntimeError("boom")):
            result = ingest_paths(job["id"], [str(bad)])
        self.assertEqual(result.get("status"), "needs_review")
        files = list_period_files(self.period["id"])
        self.assertTrue(files)
        self.assertTrue(files[0].get("parse_failed"))
        self.assertIn("could not parse", (files[0].get("classify_reason") or "").lower())
        warnings = result.get("warnings") or []
        self.assertTrue(any("GSTR-2B_July.json" in w for w in warnings), warnings)

    def test_recognised_gstr_with_zero_rows_warns(self) -> None:
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        empty = inbox / "GSTR-2B_July.json"
        empty.write_text("{", encoding="utf-8")  # broken JSON → parser returns empty
        job = start_job(self.period["id"])
        result = ingest_paths(job["id"], [str(empty)])
        files = list_period_files(self.period["id"])
        self.assertTrue(files)
        reason = (files[0].get("classify_reason") or "").lower()
        self.assertTrue(
            files[0].get("parse_failed") or "no rows" in reason or "could not parse" in reason,
            files[0],
        )
        self.assertTrue(result.get("warnings"), result)
        self.assertEqual(result.get("status"), "needs_review")

    def test_mixed_dump_keeps_bank_and_blocks_done(self) -> None:
        from make_test_dump import pdf_with_text
        from apps.engine.pipeline import get_period_pack
        from openpyxl import load_workbook

        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        (inbox / "HDFC_Statement.pdf").write_bytes(
            pdf_with_text(
                [
                    "HDFC Bank Account Statement",
                    "Account Number 50100123456789",
                    "Opening Balance 150000.00",
                    "01/07/2026 UPI merchant Withdrawal 2500.00 147500.00",
                    "Closing Balance 147500.00",
                ]
            )
        )
        (inbox / "meeting_notes.docx").write_bytes(b"PK notes")
        (inbox / "GSTR-2B_July.json").write_text("{", encoding="utf-8")
        job = start_job(self.period["id"])
        result = ingest_paths(job["id"], [str(inbox)])
        self.assertEqual(result.get("status"), "needs_review", result)
        pack = get_period_pack(self.period["id"])
        self.assertIsNotNone(pack)
        outputs = {item.get("key"): item for item in (pack.get("outputs") or [])}
        self.assertIn("bank", outputs)
        self.assertIn("needs_review", outputs)
        review_path = Path(outputs["needs_review"]["path"])
        self.assertTrue(review_path.is_file(), review_path)
        names = []
        for row in load_workbook(review_path).active.iter_rows(min_row=2, values_only=True):
            names.append(str(row[0] or ""))
        blob = " ".join(names).lower()
        self.assertIn("meeting_notes", blob)
        self.assertIn("gstr", blob)

    def test_over_cap_sets_truncation_warning(self) -> None:
        from apps.engine.dump import MAX_FOLDER_FILES, collect_inbox

        folder = Path(self._tmp.name) / "many"
        folder.mkdir()
        for i in range(MAX_FOLDER_FILES + 3):
            (folder / f"f{i}.txt").write_text("x", encoding="utf-8")
        paths, truncated = collect_inbox([str(folder)])
        self.assertEqual(len(paths), MAX_FOLDER_FILES)
        self.assertTrue(truncated)
