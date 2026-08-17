from __future__ import annotations

import ast
import os
import re
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DUMP = ROOT / "test-dump"
GSTR_2B_JSON = DUMP / "GSTR-2B_July.json"
GSTR_1_JSON = DUMP / "GSTR1_July.json"
GSTR_3B_JSON = DUMP / "GSTR3B_July.json"

GSTR_2B_XLSX = "GSTR_2B_Formatted.xlsx"
GSTR_1_XLSX = "GSTR_1_Formatted.xlsx"
GSTR_3B_XLSX = "GSTR_3B_Formatted.xlsx"

INVOICE_2B = "ACME/26-27/0142"
INVOICE_1 = "BN/101"
TAXABLE_3B = 50000

GSTR_PY = ROOT / "apps" / "engine" / "parsers" / "gstr.py"
PIPELINE_PY = ROOT / "apps" / "engine" / "pipeline.py"
DUMP_PY = ROOT / "apps" / "engine" / "dump.py"

_SCRAPE_STRINGS = ("http://", "https://")
_SCRAPE_NAMES = ("requests", "selenium", "playwright")
_RECON_HEADERS = {
    "match",
    "matched",
    "matchstatus",
    "matchresult",
    "booksref",
    "booksreference",
    "bookref",
    "recon",
    "reconstatus",
    "matchedto",
}


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pack_xlsx(pack: dict | None, filename: str, key: str) -> Path | None:
    if not pack:
        return None
    dest = Path(pack.get("path") or "") / filename
    if dest.is_file():
        return dest
    for item in pack.get("outputs") or []:
        path = Path(item.get("path") or "")
        label = str(item.get("label") or "")
        if item.get("key") == key or label.endswith(filename) or path.name == filename:
            return path if path.is_file() else dest
    return dest if dest.is_file() else None


def _sheet_names_match(names: list[str], *aliases: str) -> bool:
    wanted = {_norm_header(alias) for alias in aliases}
    return any(_norm_header(name) in wanted for name in names)


def _workbook_blob(book) -> str:
    parts: list[str] = []
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                parts.append(str(value))
    return " ".join(parts)


def _workbook_has_text(book, needle: str) -> bool:
    return needle.lower() in _workbook_blob(book).lower()


def _workbook_has_number(book, target: float, tol: float = 0.01) -> bool:
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value is None or value == "":
                    continue
                if isinstance(value, (int, float)) and abs(float(value) - target) <= tol:
                    return True
                text = str(value).replace(",", "").replace("₹", "").strip()
                try:
                    if abs(float(text) - target) <= tol:
                        return True
                except ValueError:
                    if str(int(target)) in text:
                        return True
    return False


def _recon_headers_nonempty(book) -> list[str]:
    dirty: list[str] = []
    for sheet in book.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        for index, header in enumerate(headers):
            if _norm_header(header) not in _RECON_HEADERS:
                continue
            for raw in rows[1:]:
                value = raw[index] if index < len(raw) else None
                if _cell_text(value):
                    dirty.append(f"{sheet.title}:{header}={value!r}")
                    break
    return dirty


def _source_hits(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8")
    hits: list[str] = []
    lower = text.lower()
    for token in _SCRAPE_STRINGS:
        if token in lower:
            hits.append(token)
    try:
        tree = ast.parse(text)
    except SyntaxError:
        for name in _SCRAPE_NAMES:
            if re.search(rf"\b{re.escape(name)}\b", text):
                hits.append(name)
        return hits
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                root = (alias.name or "").split(".", 1)[0]
                if root in _SCRAPE_NAMES:
                    hits.append(root)
        elif isinstance(node, ast.ImportFrom):
            root = (node.module or "").split(".", 1)[0]
            if root in _SCRAPE_NAMES:
                hits.append(root)
        elif isinstance(node, ast.Name) and node.id in _SCRAPE_NAMES:
            hits.append(node.id)
        elif isinstance(node, ast.Attribute) and node.attr in _SCRAPE_NAMES:
            hits.append(node.attr)
        elif isinstance(node, ast.Constant) and isinstance(node.value, str):
            value = node.value.lower()
            for token in _SCRAPE_STRINGS:
                if token in value:
                    hits.append(token)
    return sorted(set(hits))


def _pipeline_gstr_write_text() -> str:
    text = PIPELINE_PY.read_text(encoding="utf-8")
    keep: list[str] = []
    for line in text.splitlines():
        if re.search(r"gstr|GSTR_|parse_gstr", line):
            keep.append(line)
    return "\n".join(keep)


class _IsolatedApp(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass

    def tearDown(self) -> None:
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass
        self._tmp.cleanup()

    def _period(self):
        try:
            from apps.engine.clients import create_client
            from apps.engine.firm import save_firm
            from apps.engine.periods import create_period
        except ImportError as exc:
            self.skipTest(f"firm/clients/periods not ready: {exc}")

        save_firm("Test firm")
        client = create_client("Acme")
        return create_period(client["id"], "Jul 2026")

    def _require_dump_json(self, *paths: Path) -> None:
        missing = [path.name for path in paths if not path.is_file()]
        if missing:
            self.skipTest(f"test-dump GSTR JSON missing: {', '.join(missing)}")

    def _dump_local(self, *paths: Path) -> tuple[dict, dict]:
        try:
            from apps.engine.dump import ingest_paths, start_job
            from apps.engine.pipeline import get_period_pack
        except ImportError as exc:
            self.skipTest(f"dump/pipeline not ready: {exc}")
        try:
            from apps.engine.parsers.gstr import parse_gstr_file
        except ImportError as exc:
            self.skipTest(f"gstr parser not ready: {exc}")
        if parse_gstr_file is None:
            self.skipTest("gstr parser not ready: parse_gstr_file missing")

        self._require_dump_json(*paths)
        for path in paths:
            resolved = path.resolve()
            self.assertTrue(resolved.is_file(), path)
            self.assertTrue(
                DUMP.resolve() in resolved.parents or resolved.parent == DUMP.resolve(),
                f"dump path is not under test-dump: {resolved}",
            )

        period = self._period()
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(path) for path in paths])
        pack = get_period_pack(period["id"])
        self.assertIsNotNone(pack, "period pack was not written")
        assert pack is not None
        return period, pack


class Stage6Gstr2bGateTests(_IsolatedApp):
    def test_drop_gstr_2b_json_writes_readable_b2b_sheet(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"openpyxl not ready: {exc}")

        _period, pack = self._dump_local(GSTR_2B_JSON)
        xlsx = _pack_xlsx(pack, GSTR_2B_XLSX, "gstr_2b")
        self.assertIsNotNone(xlsx, f"missing {GSTR_2B_XLSX} under {pack.get('path')}")
        assert xlsx is not None
        self.assertTrue(xlsx.is_file(), f"missing {GSTR_2B_XLSX} under {pack.get('path')}")

        book = load_workbook(xlsx, data_only=True)
        self.assertTrue(
            _sheet_names_match(book.sheetnames, "B2B", "GSTR-2B", "GSTR 2B"),
            f"expected B2B or GSTR-2B sheet, got {book.sheetnames}",
        )
        self.assertTrue(
            _workbook_has_text(book, INVOICE_2B),
            f"{GSTR_2B_XLSX} missing invoice {INVOICE_2B}",
        )
        dirty = _recon_headers_nonempty(book)
        self.assertEqual(dirty, [], f"Match / Books ref must be empty: {dirty}")


class Stage6Gstr1GateTests(_IsolatedApp):
    def test_gstr_1_json_writes_usable_outward_register(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"openpyxl not ready: {exc}")

        _period, pack = self._dump_local(GSTR_1_JSON)
        xlsx = _pack_xlsx(pack, GSTR_1_XLSX, "gstr_1")
        self.assertIsNotNone(xlsx, f"missing {GSTR_1_XLSX} under {pack.get('path')}")
        assert xlsx is not None
        self.assertTrue(xlsx.is_file(), f"missing {GSTR_1_XLSX} under {pack.get('path')}")

        book = load_workbook(xlsx, data_only=True)
        self.assertTrue(
            _sheet_names_match(book.sheetnames, "GSTR-1", "GSTR 1", "B2B", "Outward"),
            f"expected outward / GSTR-1 sheet, got {book.sheetnames}",
        )
        self.assertTrue(
            _workbook_has_text(book, INVOICE_1),
            f"{GSTR_1_XLSX} missing invoice {INVOICE_1}",
        )
        dirty = _recon_headers_nonempty(book)
        self.assertEqual(dirty, [], f"Match / Books ref must be empty: {dirty}")


class Stage6Gstr3bGateTests(_IsolatedApp):
    def test_gstr_3b_json_writes_usable_summary(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"openpyxl not ready: {exc}")

        _period, pack = self._dump_local(GSTR_3B_JSON)
        xlsx = _pack_xlsx(pack, GSTR_3B_XLSX, "gstr_3b")
        self.assertIsNotNone(xlsx, f"missing {GSTR_3B_XLSX} under {pack.get('path')}")
        assert xlsx is not None
        self.assertTrue(xlsx.is_file(), f"missing {GSTR_3B_XLSX} under {pack.get('path')}")

        book = load_workbook(xlsx, data_only=True)
        self.assertTrue(
            _sheet_names_match(book.sheetnames, "GSTR-3B", "GSTR 3B", "Summary"),
            f"expected GSTR-3B summary sheet, got {book.sheetnames}",
        )
        self.assertTrue(
            _workbook_has_number(book, TAXABLE_3B),
            f"{GSTR_3B_XLSX} missing taxable {TAXABLE_3B}",
        )
        dirty = _recon_headers_nonempty(book)
        self.assertEqual(dirty, [], f"Match / Books ref must be empty: {dirty}")


class Stage6OfflineGateTests(_IsolatedApp):
    def test_no_gst_portal_login_or_scraping(self) -> None:
        if not GSTR_PY.is_file():
            self.skipTest("parsers/gstr.py not ready")
        if not PIPELINE_PY.is_file():
            self.skipTest("pipeline.py not ready")

        gstr_hits = _source_hits(GSTR_PY)
        self.assertEqual(gstr_hits, [], f"parsers/gstr.py must stay offline: {gstr_hits}")

        pipeline_hits = _source_hits(PIPELINE_PY)
        gstr_write = _pipeline_gstr_write_text().lower()
        write_hits = [token for token in (*_SCRAPE_STRINGS, *_SCRAPE_NAMES) if token in gstr_write]
        self.assertEqual(
            pipeline_hits,
            [],
            f"pipeline gstr write must not login/scrape: {pipeline_hits}",
        )
        self.assertEqual(write_hits, [], f"pipeline gstr write leaked scrape tokens: {write_hits}")

    def test_dump_is_fully_offline_local_test_dump(self) -> None:
        if not DUMP_PY.is_file():
            self.skipTest("dump.py not ready")
        dump_hits = _source_hits(DUMP_PY)
        self.assertEqual(dump_hits, [], f"dump.py must stay offline: {dump_hits}")

        try:
            from apps.engine.dump import list_period_files
        except ImportError as exc:
            self.skipTest(f"dump not ready: {exc}")

        sources = (GSTR_2B_JSON, GSTR_1_JSON, GSTR_3B_JSON)
        period, pack = self._dump_local(*sources)
        for path in sources:
            self.assertTrue(path.is_file())
            self.assertTrue(str(path.resolve()).lower().startswith(str(DUMP.resolve()).lower()))

        stored = list_period_files(period["id"])
        self.assertTrue(stored, "dump stored no files")
        expected = {path.name.lower() for path in sources}
        names = {str(item.get("original_name") or "").lower() for item in stored}
        self.assertTrue(expected <= names, f"expected {expected}, stored {names}")
        for item in stored:
            key = item.get("storage_key") or ""
            self.assertTrue(key, item)
            self.assertFalse(str(key).lower().startswith("http"), item)

        outputs = {Path(item.get("path") or "").name for item in (pack.get("outputs") or [])}
        self.assertTrue(
            {GSTR_2B_XLSX, GSTR_1_XLSX, GSTR_3B_XLSX} <= outputs
            or all(
                _pack_xlsx(pack, name, key) and _pack_xlsx(pack, name, key).is_file()
                for name, key in (
                    (GSTR_2B_XLSX, "gstr_2b"),
                    (GSTR_1_XLSX, "gstr_1"),
                    (GSTR_3B_XLSX, "gstr_3b"),
                )
            ),
            f"offline dump missing GSTR workbooks: {outputs}",
        )


if __name__ == "__main__":
    unittest.main()
