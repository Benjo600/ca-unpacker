from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.engine.tests.test_stage3_bank import write_hdfc
from make_test_dump import pdf_with_text


class BalanceHonestyTests(unittest.TestCase):
    def test_missing_stated_closing_is_unverified_not_match(self) -> None:
        from apps.engine.validators.balance import check_balance

        rows = [
            {"debit": 100.0, "credit": None, "balance": 900.0},
            {"debit": None, "credit": 50.0, "balance": 950.0},
        ]
        check = check_balance(rows, 1000.0, None)
        self.assertFalse(check["match"])
        self.assertEqual(check["status"], "unverified")
        self.assertIsNone(check["stated_closing"])
        self.assertNotEqual(check["stated_closing"], check["computed_closing"])

    def test_missing_opening_is_unverified_even_if_close_ties(self) -> None:
        from apps.engine.validators.balance import check_balance

        rows = [{"debit": 100.0, "credit": None, "balance": 900.0}]
        check = check_balance(rows, None, 900.0)
        self.assertTrue(check["opening_inferred"])
        self.assertFalse(check["match"])
        self.assertEqual(check["status"], "unverified")

    def test_printed_open_and_close_still_match(self) -> None:
        from apps.engine.validators.balance import check_balance

        rows = [
            {"debit": 2500.0, "credit": None, "balance": 147500.0},
            {"debit": None, "credit": 18000.0, "balance": 165500.0},
        ]
        check = check_balance(rows, 150000.0, 165500.0)
        self.assertTrue(check["match"])
        self.assertEqual(check["status"], "match")


class KeepBrokenRowsTests(unittest.TestCase):
    def test_align_keeps_row_when_running_balance_breaks(self) -> None:
        from decimal import Decimal

        from apps.engine.parsers.bank.parser import _align_to_running

        row = {
            "debit": 10.0,
            "credit": None,
            "balance": 999.0,
            "description": "UPI merchant",
            "flags": [],
        }
        kept = _align_to_running(row, Decimal("150000"))
        self.assertIsNotNone(kept)
        self.assertIn("running_balance_break", kept["flags"])

    def test_parse_keeps_corrupted_amount_line(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "HDFC_Statement_Jul2026.pdf"
            path.write_bytes(
                pdf_with_text(
                    [
                        "HDFC Bank Account Statement",
                        "Opening Balance 150000.00",
                        "01/07/2026 UPI merchant Withdrawal 2500.00 147500.00",
                        "03/07/2026 NEFT INWARD Deposit 18000.00 999999.00",
                        "05/07/2026 Cheque 112233 Withdrawal 1000.00 164500.00",
                        "Closing Balance 164500.00",
                    ]
                )
            )
            parsed = parse_bank_pdf(path)
            self.assertGreaterEqual(len(parsed["rows"]), 3)
            flagged = [
                row
                for row in parsed["rows"]
                if "running_balance_break" in (row.get("flags") or [])
            ]
            self.assertTrue(flagged, parsed["rows"])
            self.assertGreaterEqual(parsed.get("candidate_count") or 0, len(parsed["rows"]))


class BankExcelUnverifiedTests(unittest.TestCase):
    def test_cover_says_could_not_verify_not_match(self) -> None:
        from openpyxl import load_workbook

        from apps.engine.pack.bank_xlsx import write_bank_workbook
        from apps.engine.validators.balance import check_balance

        rows = [{"debit": 100.0, "credit": None, "balance": 900.0, "source": "x.pdf#p1"}]
        check = check_balance(rows, None, None)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bank.xlsx"
            write_bank_workbook(
                path,
                [
                    {
                        "rows": rows,
                        "check": check,
                        "meta": {
                            "filename": "x.pdf",
                            "profile_label": "HDFC",
                            "account_number": None,
                            "ifsc": None,
                        },
                    }
                ],
            )
            cover = load_workbook(path)["Balance Check"]
            headers = [cell.value for cell in cover[4]]
            result_col = headers.index("Result") + 1
            results = [cover.cell(row, result_col).value for row in range(5, cover.max_row + 1)]
            self.assertIn("COULD NOT VERIFY", results)
            self.assertNotIn("MATCH", results)


class AtomicReparseTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def test_failed_reparse_leaves_existing_rows(self) -> None:
        from apps.engine.clients import create_client
        from apps.engine.db import ExtractedRow, get_session
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period
        from apps.engine.pipeline import parse_period

        save_firm("Test firm")
        client = create_client("Acme")
        period = create_period(client["id"], "Jul 2026")
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        pdf = write_hdfc(inbox)
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(pdf)])

        session = get_session()
        try:
            before = session.query(ExtractedRow).filter(ExtractedRow.period_id == period["id"]).count()
        finally:
            session.close()
        self.assertGreater(before, 0)

        calls = {"n": 0}

        def boom(*_args, **_kwargs):
            calls["n"] += 1
            raise RuntimeError("forced pack failure")

        with patch("apps.engine.pipeline.write_bank_workbook", side_effect=boom):
            with self.assertRaises(RuntimeError):
                parse_period(period["id"])

        session = get_session()
        try:
            after = session.query(ExtractedRow).filter(ExtractedRow.period_id == period["id"]).count()
        finally:
            session.close()
        self.assertEqual(after, before)
