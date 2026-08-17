from __future__ import annotations

import importlib
import inspect
import os
import re
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.engine.tests.fixtures_stage5 import (
    BAD_GSTIN,
    BAD_GSTIN_INVOICE_NO,
    BAD_HSN,
    BAD_HSN_INVOICE_NO,
    write_bills_folder,
)

PURCHASE_XLSX = "Purchase_Register_Extracted.xlsx"
_CROP_MODULES = (
    "apps.engine.pipeline",
    "apps.engine.pdf_render",
    "apps.engine.crop",
    "apps.desktop.app",
)


def _import_attr(module_names: tuple[str, ...], attr: str):
    last_err: BaseException | None = None
    for name in module_names:
        try:
            module = importlib.import_module(name)
        except ImportError as exc:
            last_err = exc
            continue
        value = getattr(module, attr, None)
        if value is not None:
            return value
    if last_err is not None:
        return None
    return None


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _as_bytes(value) -> bytes:
    if value is None:
        return b""
    if isinstance(value, (bytes, bytearray)):
        return bytes(value)
    if isinstance(value, Path):
        return value.read_bytes()
    if isinstance(value, str) and Path(value).is_file():
        return Path(value).read_bytes()
    return b""


def _crop_ok_and_png(result) -> tuple[bool, bytes]:
    if isinstance(result, (bytes, bytearray)):
        return True, bytes(result)
    if isinstance(result, (str, Path)) and Path(result).is_file():
        return True, Path(result).read_bytes()
    if isinstance(result, tuple) and result:
        blob = _as_bytes(result[1]) if len(result) > 1 else b""
        return bool(result[0]), blob
    if isinstance(result, dict):
        ok = result.get("ok", True)
        for key in ("png", "png_bytes", "bytes", "image", "data", "content"):
            if result.get(key):
                return bool(ok), _as_bytes(result[key])
        if result.get("path"):
            blob = _as_bytes(result["path"])
            if blob:
                return bool(ok), blob
        url = result.get("data_url") or result.get("dataUrl")
        if isinstance(url, str) and "," in url:
            import base64

            try:
                return bool(ok), base64.b64decode(url.split(",", 1)[1])
            except Exception:
                return bool(ok), b""
        return bool(ok), b""
    return False, b""


def _call_get_source_crop(fn, context: dict):
    aliases = {
        "page": context.get("source_page"),
        "bbox": context.get("source_bbox"),
        "pdf": context.get("path"),
        "pdf_path": context.get("path"),
        "src": context.get("path"),
    }
    merged = {**aliases, **context}
    sig = inspect.signature(fn)
    kwargs = {}
    for name, param in sig.parameters.items():
        if name == "self":
            continue
        if name in merged and merged[name] is not None:
            kwargs[name] = merged[name]
        elif param.default is inspect.Parameter.empty and param.kind in (
            inspect.Parameter.POSITIONAL_ONLY,
            inspect.Parameter.POSITIONAL_OR_KEYWORD,
            inspect.Parameter.KEYWORD_ONLY,
        ):
            raise TypeError(f"get_source_crop needs {name}")
    return fn(**kwargs)


def _header_index(headers: list[str], *aliases: str) -> int | None:
    wanted = {_norm_header(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if _norm_header(header) in wanted:
            return index
    return None


def _flags_blob(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " ".join(str(item) for item in value).lower()
    return str(value).lower().replace(",", " ")


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _has_amount(value) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, (int, float)):
        return abs(float(value)) > 0
    text = str(value).strip().replace(",", "")
    if not text:
        return False
    try:
        return abs(float(text)) > 0
    except ValueError:
        return False


def _xlsx_maps(xlsx: Path) -> list[dict]:
    from openpyxl import load_workbook

    book = load_workbook(xlsx, data_only=True)
    mapped: list[dict] = []
    for sheet in book.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        gstin_i = _header_index(headers, "Supplier GSTIN", "GSTIN", "supplier_gstin")
        inv_i = _header_index(headers, "Invoice no", "Invoice number", "invoice_number", "Invoice No")
        date_i = _header_index(headers, "Date", "Invoice date", "invoice_date")
        tax_i = _header_index(headers, "Tax", "tax")
        total_i = _header_index(
            headers, "Invoice value", "invoice_value", "Total", "invoice value", "Grand total"
        )
        flags_i = _header_index(headers, "Flags", "validation_flags")
        source_i = _header_index(headers, "Source", "filename")
        for raw in rows[1:]:
            mapped.append(
                {
                    "supplier_gstin": raw[gstin_i] if gstin_i is not None and gstin_i < len(raw) else None,
                    "invoice_number": raw[inv_i] if inv_i is not None and inv_i < len(raw) else None,
                    "invoice_date": raw[date_i] if date_i is not None and date_i < len(raw) else None,
                    "tax": raw[tax_i] if tax_i is not None and tax_i < len(raw) else None,
                    "invoice_value": raw[total_i] if total_i is not None and total_i < len(raw) else None,
                    "flags": raw[flags_i] if flags_i is not None and flags_i < len(raw) else None,
                    "source": raw[source_i] if source_i is not None and source_i < len(raw) else None,
                    "_headers": headers,
                }
            )
    return mapped


def _row_matches(row: dict, needle: str) -> bool:
    blob = " ".join(_cell_text(row.get(key)) for key in ("invoice_number", "source", "supplier_gstin", "hsn"))
    return needle.lower() in blob.lower()


class _IsolatedApp(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass

    def tearDown(self) -> None:
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass
        self._tmp.cleanup()

    def _period(self):
        try:
            from apps.engine.clients import create_client
            from apps.engine.firm import save_firm
            from apps.engine.periods import create_period
        except ImportError as exc:
            self.skipTest(f"firm/clients/periods not ready: {exc}")

        save_firm("Test firm")
        client = create_client("Acme")
        return create_period(client["id"], "Jul 2026")

    def _inbox(self) -> Path:
        folder = Path(self._tmp.name) / "inbox"
        folder.mkdir(exist_ok=True)
        return folder

    def _dump_bills(self) -> tuple[dict, dict[str, Path], Path]:
        try:
            from apps.engine.dump import ingest_paths, start_job
        except ImportError as exc:
            self.skipTest(f"dump not ready: {exc}")
        inbox = self._inbox()
        try:
            paths = write_bills_folder(inbox)
        except ImportError as exc:
            self.skipTest(f"make_test_dump.pdf_with_text not ready: {exc}")
        period = self._period()
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(inbox)])
        return period, paths, inbox


class Stage5PurchasePackTests(_IsolatedApp):
    def test_dump_writes_purchase_register_xlsx(self) -> None:
        try:
            from apps.engine.pipeline import get_period_pack
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"pack/openpyxl not ready: {exc}")
        try:
            importlib.import_module("apps.engine.parsers.invoice")
        except ImportError as exc:
            self.skipTest(f"invoice parser not ready: {exc}")

        period, _paths, _inbox = self._dump_bills()
        pack = get_period_pack(period["id"])
        self.assertIsNotNone(pack, "period pack was not written")
        assert pack is not None

        xlsx = Path(pack.get("path") or "") / PURCHASE_XLSX
        if not xlsx.is_file():
            for item in pack.get("outputs") or []:
                label = str(item.get("label") or "")
                path = Path(item.get("path") or "")
                if item.get("key") == "purchase" or label.endswith(PURCHASE_XLSX) or path.name == PURCHASE_XLSX:
                    xlsx = path
                    break
        self.assertTrue(xlsx.is_file(), f"missing {PURCHASE_XLSX} under {pack.get('path')}")

        book = load_workbook(xlsx, data_only=True)
        headers: list[str] = []
        for sheet in book.worksheets:
            first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first:
                headers.extend(str(cell) if cell is not None else "" for cell in first)
        self.assertIsNotNone(_header_index(headers, "Supplier GSTIN", "GSTIN", "supplier_gstin"), headers)
        self.assertIsNotNone(
            _header_index(headers, "Invoice no", "Invoice number", "invoice_number", "Invoice No"),
            headers,
        )
        self.assertIsNotNone(_header_index(headers, "Date", "Invoice date", "invoice_date"), headers)
        self.assertIsNotNone(_header_index(headers, "Tax", "tax"), headers)
        self.assertIsNotNone(
            _header_index(headers, "Invoice value", "invoice_value", "Total", "invoice value"),
            headers,
        )

        rows = _xlsx_maps(xlsx)
        data_rows = [row for row in rows if any(_cell_text(row.get(key)) for key in ("supplier_gstin", "invoice_number", "invoice_value"))]
        self.assertTrue(data_rows, f"{PURCHASE_XLSX} has no extracted purchase rows")


class Stage5FlagTests(_IsolatedApp):
    def test_bad_gstin_flagged_checksum(self) -> None:
        try:
            from apps.engine.pipeline import get_period_pack, get_period_preview
        except ImportError as exc:
            self.skipTest(f"preview/pack not ready: {exc}")
        try:
            importlib.import_module("apps.engine.parsers.invoice")
        except ImportError as exc:
            self.skipTest(f"invoice parser not ready: {exc}")
        try:
            from apps.engine.validators.gstin import gstin_checksum_ok
        except ImportError:
            gstin_checksum_ok = None
        if gstin_checksum_ok is not None and gstin_checksum_ok(BAD_GSTIN):
            self.skipTest(f"{BAD_GSTIN} checksum unexpectedly passes")

        period, _paths, _inbox = self._dump_bills()
        preview = get_period_preview(period["id"])
        invoice_rows = [
            row
            for item in (preview.get("files") or [])
            if item.get("kind") == "invoice"
            for row in (item.get("preview") or [])
        ]
        flagged = [
            row
            for row in invoice_rows
            if BAD_GSTIN in _cell_text(row.get("supplier_gstin"))
            or BAD_GSTIN_INVOICE_NO.lower() in _cell_text(row.get("invoice_number")).lower()
            or "badgstin" in _cell_text(row.get("source")).lower()
        ]
        pack = get_period_pack(period["id"])
        xlsx_rows: list[dict] = []
        if pack:
            xlsx = Path(pack.get("path") or "") / PURCHASE_XLSX
            if not xlsx.is_file():
                for item in pack.get("outputs") or []:
                    if Path(item.get("path") or "").name == PURCHASE_XLSX:
                        xlsx = Path(item["path"])
                        break
            if xlsx.is_file():
                xlsx_rows = [
                    row
                    for row in _xlsx_maps(xlsx)
                    if BAD_GSTIN in _cell_text(row.get("supplier_gstin"))
                    or BAD_GSTIN_INVOICE_NO.lower() in _cell_text(row.get("invoice_number")).lower()
                    or "badgstin" in _cell_text(row.get("source")).lower()
                ]

        blobs = [_flags_blob(row.get("flags")) for row in flagged + xlsx_rows]
        self.assertTrue(
            flagged or xlsx_rows,
            "bad GSTIN invoice produced no ExtractedRow or Excel row",
        )
        self.assertTrue(
            any("gstin_checksum" in blob for blob in blobs),
            f"expected gstin_checksum on bad GSTIN {BAD_GSTIN}, flags={blobs}",
        )

    def test_hsn_12345_flagged_length(self) -> None:
        try:
            from apps.engine.pipeline import get_period_pack, get_period_preview
        except ImportError as exc:
            self.skipTest(f"preview/pack not ready: {exc}")
        try:
            importlib.import_module("apps.engine.parsers.invoice")
        except ImportError as exc:
            self.skipTest(f"invoice parser not ready: {exc}")

        period, _paths, _inbox = self._dump_bills()
        preview = get_period_preview(period["id"])
        invoice_rows = [
            row
            for item in (preview.get("files") or [])
            if item.get("kind") == "invoice"
            for row in (item.get("preview") or [])
        ]
        flagged = [
            row
            for row in invoice_rows
            if BAD_HSN in _cell_text(row.get("hsn"))
            or BAD_HSN_INVOICE_NO.lower() in _cell_text(row.get("invoice_number")).lower()
            or "badhsn" in _cell_text(row.get("source")).lower()
        ]
        pack = get_period_pack(period["id"])
        xlsx_rows: list[dict] = []
        if pack:
            xlsx = Path(pack.get("path") or "") / PURCHASE_XLSX
            if not xlsx.is_file():
                for item in pack.get("outputs") or []:
                    if Path(item.get("path") or "").name == PURCHASE_XLSX:
                        xlsx = Path(item["path"])
                        break
            if xlsx.is_file():
                xlsx_rows = [
                    row
                    for row in _xlsx_maps(xlsx)
                    if _row_matches(row, BAD_HSN_INVOICE_NO) or _row_matches(row, "BadHSN")
                ]
        blobs = [_flags_blob(row.get("flags")) for row in flagged + xlsx_rows]
        self.assertTrue(flagged or xlsx_rows, "HSN 12345 invoice produced no ExtractedRow or Excel row")
        self.assertTrue(
            any("hsn_length" in blob for blob in blobs),
            f"expected hsn_length on HSN {BAD_HSN}, flags={blobs}",
        )


class Stage5CropTests(_IsolatedApp):
    def test_preview_source_page_and_crop(self) -> None:
        try:
            from apps.engine.pipeline import get_period_preview
        except ImportError as exc:
            self.skipTest(f"get_period_preview not ready: {exc}")

        get_source_crop = _import_attr(_CROP_MODULES, "get_source_crop")
        try:
            importlib.import_module("apps.engine.pdf_render")
            pdf_render_ready = True
        except ImportError:
            pdf_render_ready = False

        period, _paths, _inbox = self._dump_bills()
        preview = get_period_preview(period["id"])
        files = [item for item in (preview.get("files") or []) if item.get("kind") == "invoice"]
        self.assertTrue(files, "preview has no invoice files")
        rows = [row for item in files for row in (item.get("preview") or [])]
        self.assertTrue(rows, "preview has no invoice rows")
        for row in rows:
            self.assertTrue(row.get("source_page"), row)

        if get_source_crop is None or not pdf_render_ready:
            self.skipTest("get_source_crop/pdf_render not ready")

        crop_row = next(
            (
                (item, row)
                for item in files
                for row in (item.get("preview") or [])
                if row.get("source_page")
                and (
                    row.get("source_bbox")
                    or row.get("fields")
                    or row.get("supplier_gstin")
                    or row.get("invoice_value")
                    or row.get("tax")
                    or row.get("invoice_number")
                )
            ),
            None,
        )
        if crop_row is None:
            self.skipTest("no invoice row with source_bbox or extractable fields")
        item, row = crop_row
        result = _call_get_source_crop(
            get_source_crop,
            {
                "row": row,
                "row_id": row.get("row_id"),
                "file_id": item.get("file_id"),
                "period_id": period["id"],
                "source_page": row.get("source_page"),
                "source_bbox": row.get("source_bbox") or "",
                "page": row.get("source_page"),
                "bbox": row.get("source_bbox") or "",
            },
        )
        ok, png = _crop_ok_and_png(result)
        self.assertTrue(ok, result)
        self.assertGreater(len(png), 100)


class Stage5TinyPngTests(_IsolatedApp):
    def test_tiny_png_does_not_invent_purchase_row(self) -> None:
        try:
            from apps.engine.dump import list_period_files
            from apps.engine.pipeline import get_period_pack, get_period_preview
        except ImportError as exc:
            self.skipTest(f"dump/preview not ready: {exc}")

        period, paths, _inbox = self._dump_bills()
        tiny_names = {path.name.lower() for key, path in paths.items() if path.suffix.lower() == ".png"}
        stored = list_period_files(period["id"])
        png_files = [
            item
            for item in stored
            if str(item.get("original_name") or "").lower() in tiny_names
        ]
        self.assertTrue(png_files, "1x1 PNG was not ingested")

        for item in png_files:
            kind = item.get("kind") or item.get("detected_kind")
            self.assertIn(
                kind,
                {"unknown", "invoice"},
                f"1x1 PNG should be unknown or invoice, got {kind}: {item}",
            )

        preview = get_period_preview(period["id"])
        png_preview_rows = [
            row
            for item in (preview.get("files") or [])
            if str(item.get("filename") or "").lower() in tiny_names
            for row in (item.get("preview") or [])
        ]
        for row in png_preview_rows:
            gstin = _cell_text(row.get("supplier_gstin") or row.get("gstin"))
            inv = _cell_text(row.get("invoice_number"))
            if not gstin and not inv:
                self.assertFalse(
                    _has_amount(row.get("invoice_value"))
                    or _has_amount(row.get("tax"))
                    or _has_amount(row.get("taxable_value")),
                    f"1x1 PNG invented amounts: {row}",
                )

        pack = get_period_pack(period["id"])
        if not pack:
            return
        xlsx = Path(pack.get("path") or "") / PURCHASE_XLSX
        if not xlsx.is_file():
            for item in pack.get("outputs") or []:
                if Path(item.get("path") or "").name == PURCHASE_XLSX:
                    xlsx = Path(item["path"])
                    break
        if not xlsx.is_file():
            return
        for row in _xlsx_maps(xlsx):
            source = _cell_text(row.get("source")).lower()
            from_png = Path(source).name.lower() in tiny_names or source.endswith(".png")
            if not from_png:
                continue
            gstin = _cell_text(row.get("supplier_gstin"))
            inv = _cell_text(row.get("invoice_number"))
            if not gstin and not inv:
                self.assertFalse(
                    _has_amount(row.get("invoice_value")),
                    f"purchase register invented invoice_value from 1x1 PNG: {row}",
                )


if __name__ == "__main__":
    unittest.main()
