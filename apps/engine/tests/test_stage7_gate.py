from __future__ import annotations

import ast
import os
import re
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DUMP = ROOT / "test-dump"
ZOHO_CSV = DUMP / "Zoho_Books_Invoices.csv"
TALLY_XML = DUMP / "Tally_Daybook.xml"
TALLY_ZIP = DUMP / "Tally_Backup.zip"
RANDOM_JPG = DUMP / "random_scan.jpg"

TALLY_PY = ROOT / "apps" / "engine" / "parsers" / "tally.py"
ZOHO_PY = ROOT / "apps" / "engine" / "parsers" / "zoho.py"

INV_204 = "INV-204"
PUR_88 = "PUR-88"
PUR_BOTH = "PUR-BOTH-1"
SAL_BOTH = "SAL-BOTH-1"
ZOHO_GSTIN = "27AAPFU0939F1ZV"

REGISTER_FILES = (
    ("books", "Books_Register_Extracted.xlsx"),
    ("purchase", "Purchase_Register_Extracted.xlsx"),
    ("sales", "Sales_Register_Extracted.xlsx"),
)

_BANNED_IMPORTS = {"pyodbc", "win32com", "pythoncom", "adodbapi", "requests"}
_BANNED_STRINGS = (
    "pyodbc",
    "win32com",
    "http://",
    "https://",
    "localhost:9000",
    ":9000",
)
_LIVE_API_NAMES = {
    "pyodbc",
    "win32com",
    "pythoncom",
    "adodbapi",
    "requests",
    "httpx",
    "urllib",
}


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pack_xlsx(pack: dict | None, filename: str, key: str) -> Path | None:
    if not pack:
        return None
    dest = Path(pack.get("path") or "") / filename
    if dest.is_file():
        return dest
    for item in pack.get("outputs") or []:
        path = Path(item.get("path") or "")
        label = str(item.get("label") or "")
        if item.get("key") == key or label.endswith(filename) or path.name == filename:
            return path if path.is_file() else dest
    return dest if dest.is_file() else None


def _register_paths(pack: dict | None) -> list[Path]:
    found: list[Path] = []
    seen: set[str] = set()
    for key, filename in REGISTER_FILES:
        path = _pack_xlsx(pack, filename, key)
        if path is None or not path.is_file():
            continue
        resolved = str(path.resolve()).lower()
        if resolved in seen:
            continue
        seen.add(resolved)
        found.append(path)
    if pack:
        for item in pack.get("outputs") or []:
            path = Path(item.get("path") or "")
            name = path.name.lower()
            key = str(item.get("key") or "").lower()
            if path.suffix.lower() != ".xlsx":
                continue
            if key in {"books", "purchase", "sales"} or any(
                token in name for token in ("books", "purchase", "sales")
            ):
                resolved = str(path.resolve()).lower()
                if resolved in seen or not path.is_file():
                    continue
                seen.add(resolved)
                found.append(path)
    return found


def _workbook_blob(book) -> str:
    parts: list[str] = []
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                parts.append(str(value))
    return " ".join(parts)


def _workbook_has_text(book, needle: str) -> bool:
    return needle.lower() in _workbook_blob(book).lower()


def _sheet_is(name: str, *aliases: str) -> bool:
    wanted = {_norm_header(alias) for alias in aliases}
    return _norm_header(name) in wanted


def _numeric_amounts(value) -> list[float]:
    if value is None or value == "":
        return []
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return [float(value)]
    text = str(value).replace(",", "").replace("₹", " ").replace("Rs", " ")
    found: list[float] = []
    for token in re.findall(r"-?\d+(?:\.\d+)?", text):
        try:
            found.append(float(token))
        except ValueError:
            continue
    return found


def _row_has_canonical_fields(row: dict) -> bool:
    number = _cell_text(row.get("invoice_number") or row.get("voucher_number"))
    date = _cell_text(row.get("invoice_date") or row.get("date"))
    gstin = _cell_text(row.get("supplier_gstin") or row.get("gstin"))
    amount = row.get("invoice_value")
    if amount in (None, ""):
        amount = row.get("amount")
    has_amount = False
    if isinstance(amount, (int, float)) and not isinstance(amount, bool):
        has_amount = True
    else:
        has_amount = bool(_numeric_amounts(amount))
    return bool(number and date and gstin and has_amount)


def _source_hits(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8")
    hits: list[str] = []
    lower = text.lower()
    for token in _BANNED_STRINGS:
        if token.lower() in lower:
            hits.append(token)
    try:
        tree = ast.parse(text)
    except SyntaxError:
        for name in _BANNED_IMPORTS:
            if re.search(rf"\b{re.escape(name)}\b", text):
                hits.append(name)
        return sorted(set(hits))
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                root = (alias.name or "").split(".", 1)[0]
                if root in _LIVE_API_NAMES:
                    hits.append(root)
        elif isinstance(node, ast.ImportFrom):
            root = (node.module or "").split(".", 1)[0]
            if root in _LIVE_API_NAMES:
                hits.append(root)
        elif isinstance(node, ast.Name) and node.id in _BANNED_IMPORTS:
            hits.append(node.id)
        elif isinstance(node, ast.Attribute) and node.attr in _BANNED_IMPORTS:
            hits.append(node.attr)
        elif isinstance(node, ast.Constant) and isinstance(node.value, str):
            value = node.value.lower()
            for token in _BANNED_STRINGS:
                if token.lower() in value:
                    hits.append(token)
    return sorted(set(hits))


def _both_xml() -> str:
    return f"""<?xml version="1.0"?>
<ENVELOPE>
 <HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>
 <BODY>
  <TALLYMESSAGE>
   <VOUCHER VCHTYPE="Purchase">
    <DATE>20260712</DATE>
    <VOUCHERNUMBER>{PUR_BOTH}</VOUCHERNUMBER>
    <VOUCHERTYPENAME>Purchase</VOUCHERTYPENAME>
    <PARTYLEDGERNAME>Acme Traders</PARTYLEDGERNAME>
    <PARTYGSTIN>{ZOHO_GSTIN}</PARTYGSTIN>
    <AMOUNT>-11800.00</AMOUNT>
   </VOUCHER>
   <VOUCHER VCHTYPE="Sales">
    <DATE>20260715</DATE>
    <VOUCHERNUMBER>{SAL_BOTH}</VOUCHERNUMBER>
    <VOUCHERTYPENAME>Sales</VOUCHERTYPENAME>
    <PARTYLEDGERNAME>North Retail</PARTYLEDGERNAME>
    <PARTYGSTIN>29AABCU9603R1ZM</PARTYGSTIN>
    <AMOUNT>5900.00</AMOUNT>
   </VOUCHER>
  </TALLYMESSAGE>
 </BODY>
</ENVELOPE>
"""


class _IsolatedApp(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass

    def tearDown(self) -> None:
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass
        self._tmp.cleanup()

    def _period(self):
        try:
            from apps.engine.clients import create_client
            from apps.engine.firm import save_firm
            from apps.engine.periods import create_period
        except ImportError as exc:
            self.skipTest(f"firm/clients/periods not ready: {exc}")

        save_firm("Test firm")
        client = create_client("Acme")
        return create_period(client["id"], "Jul 2026")

    def _require_dump_files(self, *paths: Path) -> None:
        missing = [path.name for path in paths if not path.is_file()]
        if missing:
            self.skipTest(f"test-dump files missing: {', '.join(missing)}")

    def _dump_paths(self, paths: list[Path]) -> tuple[dict, dict | None, dict]:
        try:
            from apps.engine.dump import ingest_paths, start_job
            from apps.engine.pipeline import get_period_pack
        except ImportError as exc:
            self.skipTest(f"dump/pipeline not ready: {exc}")

        period = self._period()
        job = start_job(period["id"])
        result = ingest_paths(job["id"], [str(path) for path in paths])
        pack = get_period_pack(period["id"])
        return period, pack, result


class Stage7ZohoGateTests(_IsolatedApp):
    def test_zoho_csv_dump_writes_register_with_inv_204(self) -> None:
        try:
            from apps.engine.parsers.zoho import parse_zoho_file
        except ImportError as exc:
            self.skipTest(f"zoho parser not ready: {exc}")
        if parse_zoho_file is None:
            self.skipTest("zoho parser not ready: parse_zoho_file missing")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"openpyxl not ready: {exc}")

        self._require_dump_files(ZOHO_CSV)
        _period, pack, job = self._dump_paths([ZOHO_CSV])
        self.assertNotEqual(job.get("status"), "failed", job)
        self.assertIsNotNone(pack, "period pack was not written")
        assert pack is not None

        registers = _register_paths(pack)
        self.assertTrue(
            registers,
            f"expected sales/books/purchase register Excel, got {pack.get('outputs')}",
        )
        found = False
        for path in registers:
            book = load_workbook(path, data_only=True)
            if _workbook_has_text(book, INV_204):
                found = True
                break
        self.assertTrue(found, f"{INV_204} missing from register workbooks {registers}")


class Stage7TallyDumpGateTests(_IsolatedApp):
    def test_tally_xml_or_zip_dump_writes_pur_88(self) -> None:
        try:
            from apps.engine.parsers.tally import parse_tally_file
        except ImportError as exc:
            self.skipTest(f"tally parser not ready: {exc}")
        if parse_tally_file is None:
            self.skipTest("tally parser not ready: parse_tally_file missing")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"openpyxl not ready: {exc}")

        sources = [path for path in (TALLY_XML, TALLY_ZIP) if path.is_file()]
        if not sources:
            self.skipTest("test-dump files missing: Tally_Daybook.xml, Tally_Backup.zip")

        _period, pack, job = self._dump_paths(sources)
        self.assertNotEqual(job.get("status"), "failed", job)
        self.assertIsNotNone(pack, "period pack was not written")
        assert pack is not None

        registers = _register_paths(pack)
        self.assertTrue(
            registers,
            f"expected purchase/books register Excel, got {pack.get('outputs')}",
        )
        found = False
        for path in registers:
            book = load_workbook(path, data_only=True)
            if _workbook_has_text(book, PUR_88):
                found = True
                break
        self.assertTrue(found, f"{PUR_88} missing from register workbooks {registers}")


class Stage7TallyBothRegistersGateTests(_IsolatedApp):
    def test_generated_purchase_and_sales_vouchers_write_both_registers(self) -> None:
        try:
            from apps.engine.parsers.tally import parse_tally_file
        except ImportError as exc:
            self.skipTest(f"tally parser not ready: {exc}")
        if parse_tally_file is None:
            self.skipTest("tally parser not ready: parse_tally_file missing")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"openpyxl not ready: {exc}")
        try:
            from apps.engine.pipeline import get_period_preview
        except ImportError:
            get_period_preview = None  # type: ignore[assignment]

        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir(exist_ok=True)
        xml_path = inbox / "Tally_Daybook_Both.xml"
        xml_path.write_text(_both_xml(), encoding="utf-8")

        period, pack, job = self._dump_paths([xml_path])
        self.assertNotEqual(job.get("status"), "failed", job)
        self.assertIsNotNone(pack, "period pack was not written")
        assert pack is not None

        registers = _register_paths(pack)
        self.assertTrue(registers, f"expected register Excel, got {pack.get('outputs')}")

        purchase_file = False
        sales_file = False
        purchase_sheet = False
        sales_sheet = False
        purchase_row = False
        sales_row = False
        type_values: list[str] = []

        for path in registers:
            name = path.name.lower()
            if "purchase" in name:
                purchase_file = True
            if "sales" in name:
                sales_file = True
            book = load_workbook(path, data_only=True)
            blob = _workbook_blob(book)
            if PUR_BOTH.lower() in blob.lower():
                purchase_row = True
            if SAL_BOTH.lower() in blob.lower():
                sales_row = True
            for sheet in book.worksheets:
                if _sheet_is(sheet.title, "Purchase", "Purchase register", "Purchases"):
                    purchase_sheet = True
                if _sheet_is(sheet.title, "Sales", "Sales register"):
                    sales_sheet = True
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [_norm_header(cell) for cell in rows[0]]
                type_idx = next(
                    (
                        index
                        for index, header in enumerate(headers)
                        if header in {"type", "vouchertype", "register", "kind"}
                    ),
                    None,
                )
                if type_idx is None:
                    continue
                for raw in rows[1:]:
                    if type_idx < len(raw):
                        type_values.append(_cell_text(raw[type_idx]).lower())

        if any("purchase" in value for value in type_values):
            purchase_row = True
        if any("sales" in value or value == "sale" for value in type_values):
            sales_row = True

        if get_period_preview is not None:
            preview = get_period_preview(period["id"], limit=20)
            for file_item in preview.get("files") or []:
                for row in file_item.get("preview") or []:
                    vtype = _cell_text(row.get("voucher_type") or row.get("register")).lower()
                    number = _cell_text(row.get("voucher_number") or row.get("invoice_number"))
                    if "purchase" in vtype or number == PUR_BOTH:
                        purchase_row = True
                    if "sales" in vtype or "sale" == vtype or number == SAL_BOTH:
                        sales_row = True

        both_files = purchase_file and sales_file
        both_sheets = purchase_sheet and sales_sheet
        both_rows = purchase_row and sales_row
        self.assertTrue(
            both_files or both_sheets or both_rows,
            "expected purchase+sales files/sheets or books rows with both registers; "
            f"files={[(p.name) for p in registers]} types={type_values}",
        )


class Stage7CanonicalFieldsGateTests(_IsolatedApp):
    def test_books_preview_or_parsed_row_has_canonical_join_fields(self) -> None:
        try:
            from apps.engine.parsers.zoho import parse_zoho_file
        except ImportError as exc:
            self.skipTest(f"zoho parser not ready: {exc}")
        if parse_zoho_file is None:
            self.skipTest("zoho parser not ready: parse_zoho_file missing")
        try:
            from apps.engine.pipeline import get_period_preview
        except ImportError:
            get_period_preview = None  # type: ignore[assignment]

        self._require_dump_files(ZOHO_CSV)
        parsed = parse_zoho_file(ZOHO_CSV, ZOHO_CSV.name)
        parsed_rows = list(parsed.get("rows") or []) if isinstance(parsed, dict) else []

        period, _pack, job = self._dump_paths([ZOHO_CSV])
        self.assertNotEqual(job.get("status"), "failed", job)

        preview_rows: list[dict] = []
        if get_period_preview is not None:
            preview = get_period_preview(period["id"], limit=20)
            for file_item in preview.get("files") or []:
                if file_item.get("kind") not in {"zoho", "tally", "books", "invoice"}:
                    continue
                preview_rows.extend(row for row in (file_item.get("preview") or []) if isinstance(row, dict))

        candidates = preview_rows + parsed_rows
        self.assertTrue(candidates, "no books/tally/zoho preview or parsed rows")
        self.assertTrue(
            any(_row_has_canonical_fields(row) for row in candidates),
            "canonical join fields missing "
            "(invoice_number/voucher_number, invoice_date/date, "
            "invoice_value/amount, supplier_gstin/gstin)",
        )


class Stage7GarbageGateTests(_IsolatedApp):
    def test_garbage_zip_unknown_or_tally_override_does_not_invent_values(self) -> None:
        try:
            from apps.engine.dump import list_period_files, override_kind, reparse_period
            from apps.engine.pipeline import get_period_preview
        except ImportError as exc:
            self.skipTest(f"dump/pipeline not ready: {exc}")
        try:
            from apps.engine.parsers.tally import parse_tally_file
        except ImportError as exc:
            self.skipTest(f"tally parser not ready: {exc}")
        if parse_tally_file is None:
            self.skipTest("tally parser not ready: parse_tally_file missing")

        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir(exist_ok=True)
        garbage = inbox / "nope_only.zip"
        with zipfile.ZipFile(garbage, "w") as archive:
            archive.writestr("nope.txt", "this is not a tally export\n")
            if RANDOM_JPG.is_file():
                archive.write(RANDOM_JPG, arcname=RANDOM_JPG.name)
            else:
                archive.writestr("random.bin", os.urandom(64))

        parsed = parse_tally_file(garbage, garbage.name)
        self.assertIsInstance(parsed, dict)
        for row in parsed.get("rows") or []:
            self._assert_no_invented_amount(row, "parse_tally_file")

        period, pack, job = self._dump_paths([garbage])
        self.assertNotEqual(job.get("status"), "failed", job)
        self.assertNotEqual(str(job.get("status") or "").lower(), "error", job)

        stored = list_period_files(period["id"])
        self.assertTrue(stored, "garbage zip was not stored")
        kinds = {str(item.get("kind") or "") for item in stored}
        self.assertTrue(
            "unknown" in kinds or "tally" in kinds,
            f"garbage zip should be unknown or tally, got {kinds}",
        )

        self._assert_no_invented_pack(pack)
        if get_period_preview is not None:
            preview = get_period_preview(period["id"], limit=50)
            for file_item in preview.get("files") or []:
                for row in file_item.get("preview") or []:
                    self._assert_no_invented_amount(row, "preview")

        target = next((item for item in stored if str(item.get("original_name") or "") == garbage.name), stored[0])
        override_kind(target["id"], "tally")
        reparsed = reparse_period(period["id"])
        self.assertNotEqual(reparsed.get("status"), "failed", reparsed)

        from apps.engine.pipeline import get_period_pack

        pack_after = get_period_pack(period["id"])
        self._assert_no_invented_pack(pack_after)
        if get_period_preview is not None:
            preview = get_period_preview(period["id"], limit=50)
            for file_item in preview.get("files") or []:
                for row in file_item.get("preview") or []:
                    self._assert_no_invented_amount(row, "tally override preview")

    def _assert_no_invented_amount(self, row: dict, where: str) -> None:
        for key in ("invoice_value", "amount"):
            value = row.get(key)
            if value in (None, "", []):
                continue
            amounts = _numeric_amounts(value)
            self.assertFalse(
                amounts,
                f"{where} invented {key}={value!r} from garbage tally input",
            )

    def _assert_no_invented_pack(self, pack: dict | None) -> None:
        if not pack:
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            return
        for path in _register_paths(pack):
            book = load_workbook(path, data_only=True)
            for sheet in book.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [_norm_header(cell) for cell in rows[0]]
                value_idxs = [
                    index
                    for index, header in enumerate(headers)
                    if header in {"invoicevalue", "amount", "value", "total"}
                ]
                if not value_idxs:
                    continue
                for raw in rows[1:]:
                    for index in value_idxs:
                        if index >= len(raw):
                            continue
                        value = raw[index]
                        if value in (None, "", "No rows extracted"):
                            continue
                        amounts = _numeric_amounts(value)
                        self.assertFalse(
                            amounts,
                            f"{path.name} invented {headers[index]}={value!r} from garbage",
                        )


class Stage7OfflineGateTests(_IsolatedApp):
    def test_tally_and_zoho_have_no_live_api(self) -> None:
        if not TALLY_PY.is_file():
            self.skipTest("parsers/tally.py not ready")
        if not ZOHO_PY.is_file():
            self.skipTest("parsers/zoho.py not ready")

        tally_hits = _source_hits(TALLY_PY)
        zoho_hits = _source_hits(ZOHO_PY)
        self.assertEqual(tally_hits, [], f"parsers/tally.py must stay offline: {tally_hits}")
        self.assertEqual(zoho_hits, [], f"parsers/zoho.py must stay offline: {zoho_hits}")


if __name__ == "__main__":
    unittest.main()
