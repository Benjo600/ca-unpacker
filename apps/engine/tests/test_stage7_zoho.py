from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.engine.tests.dump_paths import ZOHO_CSV as FIXTURE
VALID_GSTIN = "27AAPFU0939F1ZV"


class Stage7ZohoParserTests(unittest.TestCase):
    def test_fixture_csv_invoice_204(self) -> None:
        from apps.engine.parsers.zoho import parse_zoho_file

        parsed = parse_zoho_file(FIXTURE)
        self.assertEqual(len(parsed["rows"]), 1)
        row = parsed["rows"][0]
        self.assertEqual(row["invoice_number"], "INV-204")
        self.assertEqual(row["supplier_gstin"], VALID_GSTIN)
        self.assertEqual(row["invoice_value"], 5900)
        self.assertEqual(row["register"], "sales")

    def test_temp_xlsx_invoice_headers(self) -> None:
        from openpyxl import Workbook

        from apps.engine.parsers.zoho import parse_zoho_file

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "zoho_invoices.xlsx"
            book = Workbook()
            sheet = book.active
            sheet.append(["Invoice Number", "Total", "GSTIN"])
            sheet.append(["INV-1", 1180, VALID_GSTIN])
            book.save(path)
            parsed = parse_zoho_file(path, "zoho_invoices.xlsx")
        self.assertEqual(len(parsed["rows"]), 1)
        row = parsed["rows"][0]
        self.assertEqual(row["invoice_number"], "INV-1")
        self.assertEqual(row["invoice_value"], 1180)
        self.assertEqual(row["supplier_gstin"], VALID_GSTIN)
        self.assertEqual(row["register"], "sales")
        self.assertEqual(row["source"], "zoho_invoices.xlsx")

    def test_bill_number_header_is_purchase(self) -> None:
        from openpyxl import Workbook

        from apps.engine.parsers.zoho import parse_zoho_file

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "zoho_bills.xlsx"
            book = Workbook()
            sheet = book.active
            sheet.append(["Bill Number", "Vendor Name", "Total"])
            sheet.append(["BILL-9", "Acme Traders", 2500])
            book.save(path)
            parsed = parse_zoho_file(path)
        self.assertEqual(len(parsed["rows"]), 1)
        row = parsed["rows"][0]
        self.assertEqual(row["register"], "purchase")
        self.assertEqual(row["invoice_number"], "BILL-9")
        self.assertEqual(row["supplier_name"], "Acme Traders")
        self.assertEqual(row["invoice_value"], 2500)

    def test_empty_and_garbage_xlsx_are_empty_not_raised(self) -> None:
        from openpyxl import Workbook

        from apps.engine.parsers.zoho import parse_zoho_file

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            empty_path = folder / "empty.xlsx"
            Workbook().save(empty_path)
            garbage_path = folder / "garbage.xlsx"
            garbage_path.write_bytes(b"not-a-spreadsheet")
            empty = parse_zoho_file(empty_path)
            garbage = parse_zoho_file(garbage_path)
        self.assertEqual(empty["rows"], [])
        self.assertEqual(garbage["rows"], [])


if __name__ == "__main__":
    unittest.main()
