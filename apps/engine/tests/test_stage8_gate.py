from __future__ import annotations

import json
import os
import re
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.engine.tests.dump_paths import ACME, GSTR_2B as GSTR_2B_JSON, ensure_sample_dump
from apps.engine.tests.fixtures_stage5 import GOOD_GSTIN, invoice_lines, write_invoice_pdf

MASTER_XLSX = "Master_Reconciliation_Grid.xlsx"
GSTR_2B_XLSX = "GSTR_2B_Formatted.xlsx"
INVOICE_2B = "ACME/26-27/0142"
EXTRA_INV = "PORTAL-ONLY-99"
MISMATCH_INV = "MISMATCH/26-27/0007"


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pack_xlsx(pack: dict | None, filename: str, key: str) -> Path | None:
    if not pack:
        return None
    dest = Path(pack.get("path") or "") / filename
    if dest.is_file():
        return dest
    for item in pack.get("outputs") or []:
        path = Path(item.get("path") or "")
        label = str(item.get("label") or "")
        if item.get("key") == key or label.endswith(filename) or path.name == filename:
            return path if path.is_file() else dest
    return dest if dest.is_file() else None


def _sheet_named(book, *aliases: str):
    wanted = {_norm_header(alias) for alias in aliases}
    for sheet in book.worksheets:
        if _norm_header(sheet.title) in wanted:
            return sheet
    return None


def _workbook_blob(book) -> str:
    parts: list[str] = []
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                parts.append(str(value))
    return " ".join(parts)


def _grid_rows(book) -> list[dict]:
    sheet = _sheet_named(book, "Grid")
    if sheet is None:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header(cell) for cell in rows[0]]
    out: list[dict] = []
    for raw in rows[1:]:
        if not any(_cell_text(cell) for cell in raw):
            continue
        item = {}
        for index, header in enumerate(headers):
            item[header] = raw[index] if index < len(raw) else None
        out.append(item)
    return out


def _cover_count_sum(book) -> int | None:
    sheet = _sheet_named(book, "Cover")
    if sheet is None:
        return None
    keys = {
        "matched": 0,
        "booksonly": 0,
        "portalonly": 0,
        "amountmismatch": 0,
        "likely": 0,
    }
    found = False
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        label = _norm_header(row[0] if row else "")
        if label not in keys:
            continue
        found = True
        value = row[1] if len(row) > 1 else None
        try:
            keys[label] = int(value or 0)
        except (TypeError, ValueError):
            text = re.sub(r"[^\d-]", "", str(value or ""))
            keys[label] = int(text) if text else 0
    if not found:
        return None
    return sum(keys.values())


def _gstr2b_json(path: Path, invoices: list[dict]) -> Path:
    payload = {
        "gstin": "29ABCDE1234F1Z5",
        "rtnprd": "072026",
        "data": {
            "docdata": {
                "b2b": [
                    {
                        "ctin": inv.get("gstin", GOOD_GSTIN),
                        "trdnm": inv.get("trade_name", "Acme Traders"),
                        "inv": [
                            {
                                "inum": inv["inum"],
                                "dt": inv.get("dt", "12-07-2026"),
                                "val": inv["val"],
                                "txval": inv.get("txval", 10000),
                                "iamt": 0,
                                "camt": 900,
                                "samt": 900,
                            }
                        ],
                    }
                    for inv in invoices
                ]
            }
        },
    }
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path


def _recon_headers_nonempty(book) -> list[str]:
    dirty: list[str] = []
    wanted = {"match", "matched", "matchstatus", "booksref", "booksreference", "bookref"}
    for sheet in book.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        for index, header in enumerate(headers):
            if _norm_header(header) not in wanted:
                continue
            for raw in rows[1:]:
                value = raw[index] if index < len(raw) else None
                if _cell_text(value):
                    dirty.append(f"{sheet.title}:{header}={value!r}")
                    break
    return dirty


class _IsolatedApp(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass

    def tearDown(self) -> None:
        try:
            from apps.engine.db import reset_engine

            reset_engine()
        except ImportError:
            pass
        self._tmp.cleanup()

    def _period(self):
        from apps.engine.clients import create_client
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period

        save_firm("Test firm")
        client = create_client("Acme")
        return create_period(client["id"], "Jul 2026")

    def _dump_paths(self, paths: list[Path]) -> tuple[dict, dict | None, dict]:
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.pipeline import get_period_pack

        period = self._period()
        job = start_job(period["id"])
        result = ingest_paths(job["id"], [str(path) for path in paths])
        pack = get_period_pack(period["id"])
        return period, pack, result


class Stage8MasterDumpGateTests(_IsolatedApp):
    def test_acme_plus_2b_writes_master_grid_matched(self) -> None:
        from openpyxl import load_workbook

        ensure_sample_dump()
        self.assertTrue(ACME.is_file(), ACME)
        self.assertTrue(GSTR_2B_JSON.is_file(), GSTR_2B_JSON)

        _period, pack, job = self._dump_paths([ACME, GSTR_2B_JSON])
        self.assertNotEqual(job.get("status"), "failed", job)
        self.assertIsNotNone(pack, "period pack was not written")
        assert pack is not None

        xlsx = _pack_xlsx(pack, MASTER_XLSX, "master")
        self.assertIsNotNone(xlsx, f"missing {MASTER_XLSX} under {pack.get('path')}")
        assert xlsx is not None
        self.assertTrue(xlsx.is_file(), f"missing {MASTER_XLSX}")

        book = load_workbook(xlsx, data_only=True)
        blob = _workbook_blob(book)
        self.assertIn(INVOICE_2B, blob)
        grid = _grid_rows(book)
        matched = [
            row
            for row in grid
            if _cell_text(row.get("status")).lower() == "matched"
            and INVOICE_2B in f"{row.get('invoiceno2b') or ''} {row.get('invoicenobooks') or ''} {row.get('invoice2b') or ''} {row.get('invoicebooks') or ''}"
        ]
        if not matched:
            self.assertTrue(
                any(
                    _cell_text(row.get("status")).lower() == "matched" and INVOICE_2B in blob
                    for row in grid
                )
                or ("matched" in blob.lower() and INVOICE_2B in blob),
                f"expected {INVOICE_2B} as matched in {MASTER_XLSX}",
            )
        self.assertIn("recon", pack)
        self.assertEqual(pack["recon"]["counts"]["matched"], 1)


class Stage8PortalOnlyGateTests(_IsolatedApp):
    def test_synthetic_extra_2b_invoice_is_portal_only(self) -> None:
        from openpyxl import load_workbook

        ensure_sample_dump()
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir(exist_ok=True)
        bill = write_invoice_pdf(
            inbox / "Tax_Invoice_Acme.pdf",
            invoice_lines(
                invoice_no=INVOICE_2B,
                date="12/07/2026",
                taxable="10000.00",
                cgst="900.00",
                sgst="900.00",
                total="11800.00",
            ),
        )
        gstr = _gstr2b_json(
            inbox / "GSTR-2B_extra.json",
            [
                {"inum": INVOICE_2B, "val": 11800, "dt": "12-07-2026"},
                {"inum": EXTRA_INV, "val": 5900, "dt": "20-07-2026"},
            ],
        )
        _period, pack, job = self._dump_paths([bill, gstr])
        self.assertNotEqual(job.get("status"), "failed", job)
        assert pack is not None
        xlsx = _pack_xlsx(pack, MASTER_XLSX, "master")
        self.assertTrue(xlsx and xlsx.is_file(), pack)
        book = load_workbook(xlsx, data_only=True)
        blob = _workbook_blob(book)
        self.assertIn(EXTRA_INV, blob)
        self.assertIn("portal_only", blob)


class Stage8AmountMismatchGateTests(_IsolatedApp):
    def test_synthetic_amount_off_by_50_is_amount_mismatch(self) -> None:
        from openpyxl import load_workbook

        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir(exist_ok=True)
        bill = write_invoice_pdf(
            inbox / "Tax_Invoice_Mismatch.pdf",
            invoice_lines(
                invoice_no=MISMATCH_INV,
                date="12/07/2026",
                taxable="10000.00",
                cgst="900.00",
                sgst="900.00",
                total="11800.00",
            ),
        )
        gstr = _gstr2b_json(
            inbox / "GSTR-2B_mismatch.json",
            [{"inum": MISMATCH_INV, "val": 11850, "dt": "12-07-2026"}],
        )
        _period, pack, job = self._dump_paths([bill, gstr])
        self.assertNotEqual(job.get("status"), "failed", job)
        assert pack is not None
        xlsx = _pack_xlsx(pack, MASTER_XLSX, "master")
        self.assertTrue(xlsx and xlsx.is_file(), pack)
        book = load_workbook(xlsx, data_only=True)
        blob = _workbook_blob(book)
        self.assertIn("amount_mismatch", blob)
        self.assertIn(MISMATCH_INV, blob)


class Stage8TotalsGateTests(_IsolatedApp):
    def test_cover_or_grid_counts_equal_data_rows(self) -> None:
        from openpyxl import load_workbook

        ensure_sample_dump()
        _period, pack, job = self._dump_paths([ACME, GSTR_2B_JSON])
        self.assertNotEqual(job.get("status"), "failed", job)
        assert pack is not None
        xlsx = _pack_xlsx(pack, MASTER_XLSX, "master")
        self.assertTrue(xlsx and xlsx.is_file(), pack)
        book = load_workbook(xlsx, data_only=True)
        grid = _grid_rows(book)
        cover_sum = _cover_count_sum(book)
        recon = pack.get("recon") or {}
        counts = recon.get("counts") or {}
        counted = sum(int(counts.get(key) or 0) for key in ("matched", "books_only", "portal_only", "amount_mismatch", "likely"))
        self.assertEqual(counted, len(recon.get("rows") or []))
        if cover_sum is not None:
            self.assertEqual(cover_sum, len(grid))
        else:
            self.assertEqual(len(grid), counted)


class Stage8GstrOnlyNoMasterGateTests(_IsolatedApp):
    def test_gstr_only_dump_does_not_write_master_and_match_stays_empty(self) -> None:
        from openpyxl import load_workbook

        ensure_sample_dump()
        _period, pack, job = self._dump_paths([GSTR_2B_JSON])
        self.assertNotEqual(job.get("status"), "failed", job)
        assert pack is not None
        xlsx = _pack_xlsx(pack, MASTER_XLSX, "master")
        self.assertTrue(xlsx is None or not xlsx.is_file(), f"GSTR-only dump must not write {MASTER_XLSX}")
        self.assertIsNone(pack.get("recon"))
        gstr_xlsx = _pack_xlsx(pack, GSTR_2B_XLSX, "gstr_2b")
        self.assertTrue(gstr_xlsx and gstr_xlsx.is_file(), pack)
        book = load_workbook(gstr_xlsx, data_only=True)
        dirty = _recon_headers_nonempty(book)
        self.assertEqual(dirty, [], f"Match / Books ref must stay empty: {dirty}")


if __name__ == "__main__":
    unittest.main()
