from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DUMP = ROOT / "test-dump"
TALLY_XML = DUMP / "Tally_Daybook.xml"
ZOHO_CSV = DUMP / "Zoho_Books_Invoices.csv"


def _cell_texts(path: Path) -> list[str]:
    book = load_workbook(path)
    sheet = book.active
    values = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell is not None and str(cell).strip():
                values.append(str(cell))
    return values


class Stage7BooksPackTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def test_tally_and_zoho_split_into_purchase_and_sales(self) -> None:
        from apps.engine.clients import create_client
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period
        from apps.engine.pipeline import get_period_pack

        save_firm("Test firm")
        client = create_client("Acme")
        period = create_period(client["id"], "Jul 2026")
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(TALLY_XML), str(ZOHO_CSV)])
        pack = get_period_pack(period["id"])
        self.assertIsNotNone(pack)
        by_key = {item["key"]: item for item in pack["outputs"]}
        self.assertIn("books", by_key)
        self.assertIn("purchase", by_key)
        self.assertIn("sales", by_key)

        purchase_path = Path(by_key["purchase"]["path"])
        sales_path = Path(by_key["sales"]["path"])
        self.assertEqual(purchase_path.name, "Purchase_Register_Extracted.xlsx")
        self.assertEqual(sales_path.name, "Sales_Register_Extracted.xlsx")
        self.assertTrue(purchase_path.is_file())
        self.assertTrue(sales_path.is_file())

        purchase_text = " ".join(_cell_texts(purchase_path))
        self.assertTrue(
            "PUR-88" in purchase_text or "Acme" in purchase_text,
            purchase_text,
        )
        sales_text = " ".join(_cell_texts(sales_path))
        self.assertIn("INV-204", sales_text)

    def test_empty_tally_parse_writes_no_fake_row(self) -> None:
        from apps.engine.clients import create_client
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period
        from apps.engine.pipeline import get_period_pack

        garbage = Path(self._tmp.name) / "Tally_Daybook.xml"
        garbage.write_text("<ENVELOPE><TALLYMESSAGE></TALLYMESSAGE></ENVELOPE>", encoding="utf-8")

        save_firm("Test firm")
        client = create_client("Acme")
        period = create_period(client["id"], "Jul 2026")
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(garbage)])
        pack = get_period_pack(period["id"])
        self.assertTrue(pack is None or not pack.get("outputs"))


if __name__ == "__main__":
    unittest.main()
