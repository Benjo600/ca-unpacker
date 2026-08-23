from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.engine.tests.test_stage3_bank import write_hdfc
from make_test_dump import pdf_with_text


class Stage3ClassifierGateTests(unittest.TestCase):
    def test_hdfc_filename_with_bank_text_is_bank(self) -> None:
        from apps.engine.classifier import classify_from_text, classify_path

        text = (
            "HDFC Bank Account Statement\n"
            "Account Number 50100123456789\n"
            "IFSC HDFC0001234\n"
            "Opening Balance 150000.00\n"
            "Withdrawal 2500.00\n"
            "Deposit 18000.00\n"
            "Closing Balance 164500.00\n"
            "Narration UPI merchant\n"
        )
        from_text = classify_from_text("HDFC_Statement.pdf", 1, text)
        self.assertEqual(from_text.kind, "bank")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "HDFC_Statement.pdf"
            path.write_bytes(pdf_with_text(text.splitlines()))
            self.assertEqual(classify_path(path).kind, "bank")


class Stage3PackGateTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()
        from apps.engine.clients import create_client
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period

        save_firm("Test firm")
        self.client = create_client("Acme")
        self.period = create_period(self.client["id"], "Jul 2026")

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def _ingest_hdfc(self) -> dict:
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.pipeline import get_period_pack

        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir(exist_ok=True)
        pdf = write_hdfc(inbox)
        job = start_job(self.period["id"])
        ingest_paths(job["id"], [str(pdf)])
        pack = get_period_pack(self.period["id"])
        self.assertIsNotNone(pack)
        return pack  # type: ignore[return-value]

    def test_dump_writes_matching_bank_pack(self) -> None:
        pack = self._ingest_hdfc()
        self.assertTrue(pack["exists"])
        self.assertGreaterEqual(pack["row_count"], 3)
        self.assertEqual(pack["balance_status"], "match")

        xlsx = Path(pack["path"]) / "Bank_Statement_Cleaned.xlsx"
        if not xlsx.exists():
            for item in pack.get("outputs") or []:
                if item.get("key") == "bank" or str(item.get("label", "")).endswith(
                    "Bank_Statement_Cleaned.xlsx"
                ):
                    xlsx = Path(item["path"])
                    break
        self.assertTrue(xlsx.exists(), f"missing Bank_Statement_Cleaned.xlsx under {pack['path']}")

        book = load_workbook(xlsx, data_only=True)
        self.assertIn("Balance Check", book.sheetnames)

        cells: list[str] = []
        for sheet in book.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    cells.append(str(value))
        self.assertTrue(
            any("MATCH" in cell or "Match" in cell for cell in cells if "MISMATCH" not in cell.upper()),
            cells[:40],
        )

        data_sheets = [name for name in book.sheetnames if name != "Balance Check"]
        self.assertTrue(data_sheets)
        found_source = False
        found_page_ref = False
        for name in data_sheets:
            sheet = book[name]
            headers = [str(cell.value) if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            if any(header.strip().lower() == "source" for header in headers):
                found_source = True
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is not None and "#p" in str(value):
                        found_page_ref = True
        self.assertTrue(found_source, "data sheet missing Source header")
        self.assertTrue(found_page_ref, "data sheet missing #p source page marker")

    def test_period_preview_has_date_amount_and_page(self) -> None:
        from apps.engine.pipeline import get_period_preview

        self._ingest_hdfc()
        preview = get_period_preview(self.period["id"])
        files = preview.get("files") or []
        self.assertTrue(files)
        rows = [row for item in files for row in (item.get("preview") or [])]
        self.assertTrue(rows)
        for row in rows:
            self.assertTrue(row.get("date"), row)
            self.assertTrue(row.get("debit") or row.get("credit"), row)
            self.assertTrue(row.get("source_page"), row)

    def test_second_in_flight_job_if_api_rejects(self) -> None:
        from apps.engine.dump import start_job

        first = start_job(self.period["id"])
        self.assertIsNotNone(first.get("id"))
        with self.assertRaises((ValueError, RuntimeError)) as ctx:
            start_job(self.period["id"])
        self.assertIn("already", str(ctx.exception).lower())

    def test_mixed_dump_still_writes_bank_when_other_parser_raises(self) -> None:
        from unittest.mock import patch

        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.pipeline import get_period_pack

        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir(exist_ok=True)
        pdf = write_hdfc(inbox)
        bad = inbox / "GSTR1_July.json"
        bad.write_text("{not-json", encoding="utf-8")
        job = start_job(self.period["id"])
        with patch("apps.engine.pipeline.parse_gstr_file", side_effect=ValueError("bad json")):
            ingest_paths(job["id"], [str(pdf), str(bad)])
        pack = get_period_pack(self.period["id"])
        self.assertIsNotNone(pack)
        self.assertTrue(any(item.get("key") == "bank" for item in (pack.get("outputs") or [])))


if __name__ == "__main__":
    unittest.main()
