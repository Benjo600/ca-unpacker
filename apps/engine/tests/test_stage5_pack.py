from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from apps.engine.pack.table_xlsx import write_purchase_workbook, write_table
from apps.engine.pipeline import LINE_COLS, PURCHASE_COLS, get_source_crop

INVOICE_PDF = ROOT / "test-dump" / "Tax_Invoice_Acme.pdf"


def _fill_rgb(cell) -> str:
    color = getattr(cell.fill, "fgColor", None)
    rgb = getattr(color, "rgb", None) if color is not None else None
    return str(rgb or "").upper()


class PurchaseTableTests(unittest.TestCase):
    def test_write_table_includes_supplier_header(self) -> None:
        rows = [
            {
                "supplier_name": "Acme Widgets",
                "supplier_gstin": "27AAPFU0939F1ZV",
                "invoice_number": "ACME/26-27/0142",
                "invoice_date": "2026-07-12",
                "taxable_value": 10000.0,
                "tax": 1800.0,
                "invoice_value": 11800.0,
                "hsn": "998314",
                "flags": [],
                "source": "Tax_Invoice_Acme.pdf",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Purchase_Register_Extracted.xlsx"
            write_table(path, "Purchase register", rows, PURCHASE_COLS)
            book = load_workbook(path)
            sheet = book.active
            self.assertEqual(sheet.title, "Purchase register")
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("Supplier", headers)
            self.assertEqual(headers[0], "Supplier")
            self.assertEqual(sheet["A2"].value, "Acme Widgets")
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertTrue(sheet.auto_filter.ref)

    def test_flag_row_fill_when_gstin_checksum(self) -> None:
        rows = [
            {
                "supplier_name": "Bad GSTIN Co",
                "supplier_gstin": "27AAAAA0000A1Z5",
                "invoice_number": "INV-1",
                "invoice_date": "2026-07-01",
                "taxable_value": 100.0,
                "tax": 18.0,
                "invoice_value": 118.0,
                "hsn": "9983",
                "flags": ["gstin_checksum"],
                "source": "bad.pdf",
            },
            {
                "supplier_name": "Clean Co",
                "supplier_gstin": "27AAPFU0939F1ZV",
                "invoice_number": "INV-2",
                "invoice_date": "2026-07-02",
                "taxable_value": 200.0,
                "tax": 36.0,
                "invoice_value": 236.0,
                "hsn": "9983",
                "flags": [],
                "source": "ok.pdf",
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "purchase.xlsx"
            write_table(path, "Purchase register", rows, PURCHASE_COLS)
            sheet = load_workbook(path).active
            flagged = [_fill_rgb(cell) for cell in sheet[2]]
            clean = [_fill_rgb(cell) for cell in sheet[3]]
            self.assertTrue(any("FFCDD2" in value for value in flagged), flagged)
            self.assertFalse(any("FFCDD2" in value for value in clean), clean)
            tax_col = [cell.value for cell in sheet[1]].index("Taxable") + 1
            self.assertEqual(sheet.cell(2, tax_col).number_format, "#,##0.00")

    def test_purchase_workbook_has_invoice_and_line_item_sheets(self) -> None:
        invoices = [
            {
                "supplier_name": "Acme Widgets",
                "supplier_gstin": "27AAPFU0939F1ZV",
                "invoice_number": "ACME/26-27/0142",
                "invoice_date": "2026-07-12",
                "taxable_value": 10000.0,
                "tax": 1800.0,
                "invoice_value": 11800.0,
                "hsn": "998314",
                "flags": [],
                "source": "Tax_Invoice_Acme.pdf",
            }
        ]
        items = [
            {
                "invoice_number": "ACME/26-27/0142",
                "description": "Consulting",
                "hsn": "998314",
                "qty": 1,
                "rate": 7000.0,
                "taxable": 7000.0,
                "tax": None,
                "amount": 8260.0,
                "source": "Tax_Invoice_Acme.pdf",
            },
            {
                "invoice_number": "ACME/26-27/0142",
                "description": "Software",
                "hsn": "997331",
                "qty": 1,
                "rate": 3000.0,
                "taxable": 3000.0,
                "tax": None,
                "amount": 3540.0,
                "source": "Tax_Invoice_Acme.pdf",
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Purchase_Register_Extracted.xlsx"
            write_purchase_workbook(path, invoices, items, PURCHASE_COLS, LINE_COLS)
            book = load_workbook(path)
            titles = [sheet.title for sheet in book.worksheets]
            self.assertIn("Invoices", titles)
            self.assertIn("Line items", titles)
            lines = book["Line items"]
            headers = [cell.value for cell in lines[1]]
            self.assertIn("HSN", headers)
            self.assertEqual(lines.max_row, 3)


class InvoiceCropTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def test_get_source_crop_dumped_tax_invoice(self) -> None:
        from apps.engine.clients import create_client
        from apps.engine.dump import ingest_paths, list_period_files, start_job
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period

        self.assertTrue(INVOICE_PDF.is_file(), INVOICE_PDF)
        save_firm("Test firm")
        client = create_client("Acme")
        period = create_period(client["id"], "Jul 2026")
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(INVOICE_PDF)])
        files = list_period_files(period["id"])
        stored = next(
            (item for item in files if "Tax_Invoice" in (item.get("original_name") or "")),
            files[0] if files else None,
        )
        self.assertIsNotNone(stored)
        result = get_source_crop(stored["id"], 1, "")
        self.assertTrue(result.get("ok"), result)
        png = b""
        if result.get("path") and Path(result["path"]).is_file():
            png = Path(result["path"]).read_bytes()
        elif isinstance(result.get("data_url"), str) and "," in result["data_url"]:
            import base64

            png = base64.b64decode(result["data_url"].split(",", 1)[1])
        self.assertGreater(len(png), 100)
        self.assertTrue(png.startswith(b"\x89PNG\r\n\x1a\n"))


class ImageCropTests(unittest.TestCase):
    def test_crop_image_png_with_bbox(self) -> None:
        from apps.engine.pdf_render import crop_image_png

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "invoice.png"
            image = Image.new("RGB", (200, 300), (255, 255, 255))
            for x in range(20, 80):
                for y in range(220, 260):
                    image.putpixel((x, y), (20, 80, 180))
            image.save(src)
            dest = Path(tmp) / "crop.png"
            # Bottom-left origin: pixel_top = 300 - (40 + 40) = 220 (blue block).
            out = crop_image_png(src, "20.0,40.0,60.0,40.0", dest, pad=0.0)
            self.assertEqual(out, dest)
            self.assertTrue(dest.is_file())
            self.assertGreater(dest.stat().st_size, 50)
            with Image.open(dest) as cropped:
                cropped.load()
                self.assertEqual(cropped.format, "PNG")
                self.assertEqual(cropped.size, (60, 40))
                self.assertEqual(cropped.getpixel((5, 5)), (20, 80, 180))


if __name__ == "__main__":
    unittest.main()
