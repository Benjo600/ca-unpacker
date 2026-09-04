from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from make_test_dump import pdf_with_text


def write_hdfc(folder: Path) -> Path:
    path = folder / "HDFC_Statement_Jul2026.pdf"
    path.write_bytes(
        pdf_with_text(
            [
                "HDFC Bank Account Statement",
                "Account Number 50100123456789",
                "IFSC HDFC0001234",
                "Opening Balance 150000.00",
                "01/07/2026 UPI merchant Withdrawal 2500.00 147500.00",
                "03/07/2026 NEFT INWARD Deposit 18000.00 165500.00",
                "05/07/2026 Cheque 112233 Withdrawal 1000.00 164500.00",
                "Closing Balance 164500.00",
            ]
        )
    )
    return path


class BankParserTests(unittest.TestCase):
    def test_parse_hdfc_dummy(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        with tempfile.TemporaryDirectory() as tmp:
            path = write_hdfc(Path(tmp))
            parsed = parse_bank_pdf(path)
            self.assertEqual(parsed["profile"], "hdfc")
            self.assertGreaterEqual(len(parsed["rows"]), 3)
            dates = [row["date"] for row in parsed["rows"]]
            self.assertIn("2026-07-01", dates)
            check = check_balance(parsed["rows"], parsed["opening_balance"], parsed["stated_closing"])
            self.assertTrue(check["match"], check)


class MonthNameDateTests(unittest.TestCase):
    def test_kotak_style_dates_parse_and_balance(self) -> None:
        from apps.engine.parsers.bank.money import normalize_date
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        self.assertEqual(normalize_date("05 Jul 2026"), "2026-07-05")
        self.assertEqual(normalize_date("14 Jul2026"), "2026-07-14")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Kotak_Statement_Jul2026.pdf"
            path.write_bytes(
                pdf_with_text(
                    [
                        "Kotak Mahindra Bank Account Statement",
                        "Opening Balance 11.71",
                        "05 Jul 2026 UPI inward Deposit 4000.00 4011.71",
                        "05 Jul 2026 UPI merchant Withdrawal 3800.00 211.71",
                        "Closing Balance 211.71",
                    ]
                )
            )
            parsed = parse_bank_pdf(path)
            self.assertEqual(parsed["profile"], "kotak")
            self.assertGreaterEqual(len(parsed["rows"]), 2, parsed["rows"])
            dates = [row["date"] for row in parsed["rows"]]
            self.assertIn("2026-07-05", dates)
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertEqual(check["status"], "match", check)


class MarkdownTableTests(unittest.TestCase):
    def test_kotak_markdown_table_maps_debit_credit(self) -> None:
        from apps.engine.parsers.bank.profiles import detect_profile
        from apps.engine.parsers.bank.tables import rows_from_markdown
        from apps.engine.validators.balance import check_balance

        markdown = """
| Date | Description | Chq/Ref. No. | Withdrawal (Dr.) | Deposit (Cr.) | Balance |
| --- | --- | --- | --- | --- | --- |
| - | Opening Balance | - | - | - | 11.71 |
| 05 Jul 2026 | UPI inward | UPI-1 |  | 4,000.00 | 4,011.71 |
| 05 Jul 2026 | UPI merchant | UPI-2 | 3,800.00 |  | 211.71 |
"""
        profile = detect_profile("Kotak Mahindra Bank", "kotak.pdf")
        rows = rows_from_markdown(markdown, profile, "kotak.pdf")
        self.assertEqual(len(rows), 2, rows)
        self.assertEqual(rows[0]["credit"], 4000.0)
        self.assertEqual(rows[1]["debit"], 3800.0)
        check = check_balance(rows, 11.71, 211.71)
        self.assertEqual(check["status"], "match", check)

    def test_dateless_amount_row_is_its_own_transaction(self) -> None:
        from apps.engine.parsers.bank.profiles import detect_profile
        from apps.engine.parsers.bank.tables import rows_from_markdown
        from apps.engine.validators.balance import check_balance

        markdown = """
| Date | Description | Chq/Ref. No. | Withdrawal (Dr.) | Deposit (Cr.) | Balance |
| --- | --- | --- | --- | --- | --- |
| 14 Jul 2026 | UPI wrap debit | UPI-21 | 100.00 |  | 65.82 |
|  | Payment from previous page | UPI-22 |  | 200.00 | 265.82 |
| 14 Jul 2026 | UPI cake | UPI-23 | 140.00 |  | 125.82 |
"""
        profile = detect_profile("Kotak Mahindra Bank", "kotak.pdf")
        rows = rows_from_markdown(markdown, profile, "kotak.pdf")
        self.assertEqual(len(rows), 3, rows)
        self.assertEqual([row["date"] for row in rows], ["2026-07-14"] * 3)
        self.assertEqual(rows[1]["credit"], 200.0)
        self.assertEqual(rows[1]["balance"], 265.82)
        self.assertEqual(rows[2]["debit"], 140.0)
        self.assertEqual(rows[2]["balance"], 125.82)
        check = check_balance(rows, 165.82, 125.82)
        self.assertEqual(check["status"], "match", check)

    def test_date_inside_narration_does_not_start_a_new_row(self) -> None:
        from apps.engine.parsers.bank.profiles import detect_profile
        from apps.engine.parsers.bank.tables import rows_from_markdown

        markdown = """
| Date | Description | Chq/Ref. No. | Withdrawal (Dr.) | Deposit (Cr.) | Balance |
| --- | --- | --- | --- | --- | --- |
| 30 Jul 2026 | CHRG: DCC FEE FOR 2877 ECOM TXN |  | 1.77 |  | 1,579.85 |
|  | ON 12-JUN-2026 |  |  |  |  |
| 31 Jul 2026 | UPI next | UPI-1 |  | 100.00 | 1,679.85 |
"""
        profile = detect_profile("Kotak Mahindra Bank", "kotak.pdf")
        rows = rows_from_markdown(markdown, profile, "kotak.pdf")
        self.assertEqual(len(rows), 2, rows)
        self.assertEqual(rows[0]["date"], "2026-07-30")
        self.assertEqual(rows[0]["debit"], 1.77)
        self.assertIn("12-JUN-2026", rows[0]["description"])
        self.assertEqual(rows[1]["date"], "2026-07-31")


class RunningGapRepairTests(unittest.TestCase):
    def test_inserts_missing_credit_so_later_dates_stay_put(self) -> None:
        from decimal import Decimal

        from apps.engine.parsers.bank.parser import _finish_rows
        from apps.engine.parsers.bank.profiles import detect_profile
        from apps.engine.validators.balance import check_balance

        profile = detect_profile("Kotak Mahindra Bank", "kotak.pdf")
        raw = [
            {
                "date": "2026-07-14",
                "description": "UPI wrap debit",
                "debit": 100.0,
                "credit": None,
                "balance": 65.82,
                "flags": [],
            },
            {
                "date": "2026-07-14",
                "description": "UPI cake",
                "debit": 140.0,
                "credit": None,
                "balance": 125.82,
                "flags": [],
            },
            {
                "date": "2026-07-15",
                "description": "UPI next day",
                "debit": None,
                "credit": 100.0,
                "balance": 225.82,
                "flags": [],
            },
        ]
        rows = _finish_rows(raw, Decimal("165.82"), None, None, profile)
        self.assertEqual(len(rows), 4, rows)
        self.assertEqual(rows[1]["credit"], 200.0)
        self.assertEqual(rows[1]["balance"], 265.82)
        self.assertEqual(rows[2]["date"], "2026-07-14")
        self.assertEqual(rows[2]["debit"], 140.0)
        self.assertEqual(rows[3]["date"], "2026-07-15")
        check = check_balance(rows, 165.82, 225.82)
        self.assertEqual(check["status"], "match", check)


class SplitLayoutTests(unittest.TestCase):
    def test_date_on_one_line_amounts_on_next(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "HDFC_Statement_Jul2026.pdf"
            path.write_bytes(
                pdf_with_text(
                    [
                        "HDFC Bank Account Statement",
                        "Opening Balance 150000.00",
                        "01/07/2026 UPI merchant Withdrawal",
                        "2500.00 147500.00",
                        "03/07/2026 NEFT INWARD Deposit",
                        "18000.00 165500.00",
                        "05/07/2026 Cheque 112233 Withdrawal",
                        "1000.00 164500.00",
                        "Closing Balance 164500.00",
                    ]
                )
            )
            parsed = parse_bank_pdf(path)
            self.assertGreaterEqual(len(parsed["rows"]), 3, parsed["rows"])
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertEqual(check["status"], "match", check)

    def test_wrapped_narration_then_amounts(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "HDFC_Statement_Jul2026.pdf"
            path.write_bytes(
                pdf_with_text(
                    [
                        "HDFC Bank Account Statement",
                        "Opening Balance 150000.00",
                        "01/07/2026 UPI-MEHTA TRADING PVT LTD",
                        "INVOICE JULY MUMBAI",
                        "2500.00 147500.00",
                        "Closing Balance 147500.00",
                    ]
                )
            )
            parsed = parse_bank_pdf(path)
            self.assertEqual(len(parsed["rows"]), 1, parsed["rows"])
            row = parsed["rows"][0]
            self.assertEqual(row["debit"], 2500.0)
            self.assertIn("MEHTA", row["description"].upper())
            self.assertIn("MUMBAI", row["description"].upper())
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertEqual(check["status"], "match", check)

    def test_column_header_debit_credit_layout(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "HDFC_Statement_Jul2026.pdf"
            path.write_bytes(
                pdf_with_xy(
                    [
                        (40, 800, "HDFC Bank Account Statement"),
                        (40, 780, "Opening Balance 150000.00"),
                        (40, 750, "Txn Date"),
                        (120, 750, "Narration"),
                        (320, 750, "Withdrawal"),
                        (400, 750, "Deposit"),
                        (480, 750, "Balance"),
                        (40, 720, "01/07/2026"),
                        (120, 720, "UPI merchant"),
                        (320, 720, "2500.00"),
                        (480, 720, "147500.00"),
                        (40, 700, "03/07/2026"),
                        (120, 700, "NEFT INWARD"),
                        (400, 700, "18000.00"),
                        (480, 700, "165500.00"),
                        (40, 670, "Closing Balance 165500.00"),
                    ]
                )
            )
            parsed = parse_bank_pdf(path)
            self.assertGreaterEqual(len(parsed["rows"]), 2, parsed)
            dates = [row["date"] for row in parsed["rows"]]
            self.assertIn("2026-07-01", dates)
            self.assertIn("2026-07-03", dates)
            first = next(row for row in parsed["rows"] if row["date"] == "2026-07-01")
            self.assertEqual(first["debit"], 2500.0)
            self.assertTrue(first.get("credit") in (None, 0.0))
            second = next(row for row in parsed["rows"] if row["date"] == "2026-07-03")
            self.assertEqual(second["credit"], 18000.0)
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertEqual(check["status"], "match", (check, parsed["rows"]))
            self.assertIn(parsed.get("parse_strategy"), {"pdfplumber", "columns", "lines"})


class PdfPlumberTableTests(unittest.TestCase):
    def test_extract_tables_from_column_layout(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.parsers.bank.plumber import rows_from_pdfplumber
        from apps.engine.parsers.bank.profiles import detect_profile
        from apps.engine.validators.balance import check_balance

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "HDFC_Statement_Jul2026.pdf"
            path.write_bytes(
                pdf_with_xy(
                    [
                        (40, 800, "HDFC Bank Account Statement"),
                        (40, 780, "Opening Balance 150000.00"),
                        (40, 750, "Txn Date"),
                        (120, 750, "Narration"),
                        (320, 750, "Withdrawal"),
                        (400, 750, "Deposit"),
                        (480, 750, "Balance"),
                        (40, 720, "01/07/2026"),
                        (120, 720, "UPI merchant"),
                        (320, 720, "2500.00"),
                        (480, 720, "147500.00"),
                        (40, 700, "03/07/2026"),
                        (120, 700, "NEFT INWARD"),
                        (400, 700, "18000.00"),
                        (480, 700, "165500.00"),
                        (40, 670, "Closing Balance 165500.00"),
                    ]
                )
            )
            profile = detect_profile("HDFC Bank Account Statement", path.name)
            rows = rows_from_pdfplumber(path, profile, path.name)
            self.assertGreaterEqual(len(rows), 2, rows)
            parsed = parse_bank_pdf(path)
            self.assertEqual(parsed["parse_strategy"], "pdfplumber", parsed)
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertEqual(check["status"], "match", check)


def pdf_with_xy(items: list[tuple[float, float, str]]) -> bytes:
    content = "BT /F1 10 Tf\n"
    for x, y, text in items:
        safe = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        content += f"1 0 0 1 {x:.1f} {y:.1f} Tm ({safe}) Tj\n"
    content += "ET\n"
    stream = content.encode("latin-1", errors="replace")
    objects = [
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n",
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n",
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 595 842]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj\n",
        f"4 0 obj<</Length {len(stream)}>>stream\n".encode("ascii") + stream + b"endstream\nendobj\n",
        b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n",
    ]
    body = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objects:
        offsets.append(len(body))
        body += obj
    xref_pos = len(body)
    xref = f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n"
    for off in offsets[1:]:
        xref += f"{off:010d} 00000 n \n"
    trailer = (
        f"trailer<</Size {len(objects)+1}/Root 1 0 R>>\nstartxref\n{xref_pos}\n%%EOF\n"
    )
    return body + xref.encode("ascii") + trailer.encode("ascii")


class BankPackPipelineTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def test_dump_writes_excel(self) -> None:
        from apps.engine.clients import create_client
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period
        from apps.engine.pipeline import get_period_pack

        save_firm("Test firm")
        client = create_client("Acme")
        period = create_period(client["id"], "Jul 2026")
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        write_hdfc(inbox)
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(inbox / "HDFC_Statement_Jul2026.pdf")])
        pack = get_period_pack(period["id"])
        self.assertIsNotNone(pack)
        self.assertTrue(pack["exists"])
        self.assertGreaterEqual(pack["row_count"], 3)
        self.assertEqual(pack["balance_status"], "match")
        self.assertTrue(Path(pack["path"]).exists())


class ComplexStatementTests(unittest.TestCase):
    def test_complex_folder_balances(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        from apps.engine.tests.dump_paths import BANKS_MESSY

        folder = BANKS_MESSY
        files = list(folder.glob("*_complex.pdf"))
        self.assertGreaterEqual(len(files), 3)
        for path in files:
            parsed = parse_bank_pdf(path)
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertGreaterEqual(len(parsed["rows"]), 10, path.name)
            self.assertEqual(check["status"], "match", (path.name, check))


class BankWorkbookPolishTests(unittest.TestCase):
    def test_single_file_sheet_cover_source_and_flags(self) -> None:
        from openpyxl import load_workbook

        from apps.engine.pack.bank_xlsx import write_bank_workbook
        from apps.engine.validators.balance import check_balance

        rows = [
            {
                "date": "2026-07-01",
                "description": "UPI merchant",
                "cheque_ref": None,
                "debit": 2500.00,
                "credit": None,
                "balance": 147500.00,
                "source": "HDFC_Statement_Jul2026.pdf#p1",
            },
            {
                "date": "2026-07-03",
                "description": "NEFT INWARD",
                "cheque_ref": None,
                "debit": None,
                "credit": 18000.00,
                "balance": 165500.00,
                "source": "HDFC_Statement_Jul2026.pdf#p1",
            },
        ]
        check = check_balance(rows, 150000.00, 165500.00)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bank.xlsx"
            write_bank_workbook(
                path,
                [
                    {
                        "rows": rows,
                        "check": check,
                        "meta": {
                            "filename": "HDFC_Statement_Jul2026.pdf",
                            "profile_label": "HDFC",
                            "account_number": "50100123456789",
                            "ifsc": "HDFC0001234",
                        },
                    }
                ],
            )
            book = load_workbook(path)
            self.assertIn("Transactions", book.sheetnames)
            cover = book["Balance Check"]
            headers = [cell.value for cell in cover[4]]
            self.assertIn("Result", headers)
            result_col = headers.index("Result") + 1
            results = [cover.cell(row, result_col).value for row in range(5, cover.max_row + 1)]
            self.assertTrue(any(value in {"MATCH", "MISMATCH"} for value in results), results)
            sheet = book["Transactions"]
            tx_headers = [cell.value for cell in sheet[1]]
            self.assertIn("Flags", tx_headers)
            self.assertIn("Source", tx_headers)
            source_col = tx_headers.index("Source") + 1
            sources = [sheet.cell(row, source_col).value for row in range(2, sheet.max_row + 1)]
            self.assertTrue(any(source and "#p" in str(source) for source in sources), sources)

    def test_opening_inferred_when_opening_missing(self) -> None:
        from apps.engine.validators.balance import check_balance

        rows = [{"debit": 100.0, "credit": None, "balance": 900.0}]
        check = check_balance(rows, None, 900.0)
        self.assertIn("opening_inferred", check)
        self.assertTrue(check["opening_inferred"])


class DumpJobTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def _period(self):
        from apps.engine.clients import create_client
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period

        save_firm("Test firm")
        client = create_client("Acme")
        return create_period(client["id"], "Jul 2026")

    def test_start_job_rejects_in_progress(self) -> None:
        from apps.engine.db import Job, get_session
        from apps.engine.dump import start_job

        period = self._period()
        first = start_job(period["id"])
        with self.assertRaises(ValueError) as ctx:
            start_job(period["id"])
        self.assertIn("already", str(ctx.exception).lower())

        session = get_session()
        try:
            row = session.get(Job, first["id"])
            self.assertIsNotNone(row)
            row.status = "done"
            session.commit()
        finally:
            session.close()

        second = start_job(period["id"])
        self.assertNotEqual(second["id"], first["id"])

    def test_ingest_failure_marks_job_failed_without_traceback(self) -> None:
        from unittest.mock import patch

        from apps.engine.dump import get_job, ingest_paths, start_job

        period = self._period()
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        source = inbox / "notes.txt"
        source.write_text("hello", encoding="utf-8")
        job = start_job(period["id"])
        with patch(
            "apps.engine.dump.parse_period_banks",
            side_effect=RuntimeError("Traceback (most recent call last):\n  File boom"),
        ):
            with self.assertRaises(RuntimeError):
                ingest_paths(job["id"], [str(source)])
        done = get_job(job["id"])
        self.assertIsNotNone(done)
        self.assertEqual(done["status"], "failed")
        self.assertNotIn("Traceback", done["error_message"] or "")
        self.assertTrue(done["error_message"])

    def test_reparse_period_after_override(self) -> None:
        from apps.engine.dump import (
            ingest_paths,
            list_period_files,
            override_kind,
            reparse_period,
            start_job,
        )
        from apps.engine.pipeline import get_period_pack

        period = self._period()
        inbox = Path(self._tmp.name) / "inbox"
        inbox.mkdir()
        write_hdfc(inbox)
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(inbox / "HDFC_Statement_Jul2026.pdf")])
        files = list_period_files(period["id"])
        bank = next(row for row in files if row["kind"] == "bank")
        updated = override_kind(bank["id"], "bank")
        self.assertEqual(updated["kind"], "bank")
        parsed = reparse_period(period["id"])
        self.assertEqual(parsed["status"], "done")
        pack = get_period_pack(period["id"])
        self.assertIsNotNone(pack)
        self.assertTrue(pack["exists"])


class LongStatementTests(unittest.TestCase):
    def test_long_folder_if_present(self) -> None:
        from apps.engine.parsers.bank.parser import parse_bank_pdf
        from apps.engine.validators.balance import check_balance

        from apps.engine.tests.dump_paths import BANKS_MESSY

        folder = BANKS_MESSY
        files = list(folder.glob("*_LONG.pdf"))
        if not files:
            self.skipTest("long statements not generated yet")
        for path in files:
            parsed = parse_bank_pdf(path)
            check = check_balance(
                parsed["rows"], parsed["opening_balance"], parsed["stated_closing"]
            )
            self.assertGreaterEqual(len(parsed["rows"]), 80, path.name)
            self.assertEqual(check["status"], "match", (path.name, check, len(parsed["rows"])))


if __name__ == "__main__":
    unittest.main()
