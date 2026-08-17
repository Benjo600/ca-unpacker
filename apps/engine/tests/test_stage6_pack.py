from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DUMP = ROOT / "test-dump"
GSTR_2B_JSON = DUMP / "GSTR-2B_July.json"
GSTR_1_JSON = DUMP / "GSTR1_July.json"
GSTR_3B_JSON = DUMP / "GSTR3B_July.json"


def _headers(sheet) -> list:
    return [cell.value for cell in sheet[1]]


def _col(sheet, name: str) -> int:
    return _headers(sheet).index(name) + 1


class GstrWorkbookWriteTests(unittest.TestCase):
    def test_2b_b2b_has_empty_match_and_acme_invoice(self) -> None:
        from apps.engine.pack.gstr_xlsx import write_gstr_2b
        from apps.engine.parsers.gstr import parse_gstr_file

        parsed = parse_gstr_file(GSTR_2B_JSON, "gstr_2b")
        rows = [dict(row, match_status="matched", books_ref="PUR-88") for row in parsed["rows"]]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "GSTR_2B_Formatted.xlsx"
            write_gstr_2b(
                path,
                rows,
                {"gstin": parsed.get("gstin"), "period": parsed.get("period")},
            )
            book = load_workbook(path)
            self.assertIn("B2B", book.sheetnames)
            self.assertIn("Cover", book.sheetnames)
            self.assertIn("Flags", book.sheetnames)
            self.assertNotIn("CDN", book.sheetnames)
            sheet = book["B2B"]
            headers = _headers(sheet)
            self.assertIn("Match", headers)
            self.assertIn("Books ref", headers)
            invoices = [sheet.cell(row, _col(sheet, "Invoice no")).value for row in range(2, sheet.max_row + 1)]
            self.assertTrue(any(value and "ACME" in str(value) for value in invoices), invoices)
            match_col = _col(sheet, "Match")
            books_col = _col(sheet, "Books ref")
            for excel_row in range(2, sheet.max_row + 1):
                self.assertIn(sheet.cell(excel_row, match_col).value, (None, ""))
                self.assertIn(sheet.cell(excel_row, books_col).value, (None, ""))
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertTrue(sheet.auto_filter.ref)
            tax_col = _col(sheet, "Taxable")
            self.assertEqual(sheet.cell(2, tax_col).number_format, "#,##0.00")
            self.assertEqual(book["Flags"]["A2"].value, "none")

    def test_gstr1_writes_hsn_sheet(self) -> None:
        from apps.engine.pack.gstr_xlsx import write_gstr_1
        from apps.engine.parsers.gstr import parse_gstr_file

        parsed = parse_gstr_file(GSTR_1_JSON, "gstr_1")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "GSTR_1_Formatted.xlsx"
            write_gstr_1(path, parsed["rows"], {"gstin": parsed.get("gstin"), "period": parsed.get("period")})
            book = load_workbook(path)
            self.assertIn("HSN", book.sheetnames)
            self.assertIn("B2B", book.sheetnames)
            hsn_values = [cell.value for cell in book["HSN"]["A"]]
            self.assertIn("9983", hsn_values)
            match_col = _col(book["B2B"], "Match")
            self.assertIn(book["B2B"].cell(2, match_col).value, (None, ""))

    def test_3b_summary_has_50000_taxable(self) -> None:
        from apps.engine.pack.gstr_xlsx import write_gstr_3b
        from apps.engine.parsers.gstr import parse_gstr_file

        parsed = parse_gstr_file(GSTR_3B_JSON, "gstr_3b")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "GSTR_3B_Formatted.xlsx"
            write_gstr_3b(path, parsed["rows"], {"gstin": parsed.get("gstin"), "period": parsed.get("period")})
            book = load_workbook(path)
            self.assertIn("Summary", book.sheetnames)
            self.assertNotIn("Match", _headers(book["Summary"]))
            taxables = [cell.value for cell in book["Summary"]["B"]]
            self.assertIn(50000, taxables)
            tax_col = _col(book["Summary"], "Taxable")
            self.assertEqual(book["Summary"].cell(2, tax_col).number_format, "#,##0.00")

    def test_2b_cdn_sheet_only_when_present(self) -> None:
        from apps.engine.pack.gstr_xlsx import write_gstr_2b

        rows = [
            {
                "gstin": "27AAPFU0939F1ZV",
                "trade_name": "Acme Traders",
                "invoice_number": "CN-1",
                "invoice_date": "15-07-2026",
                "invoice_value": 118.0,
                "taxable": 100.0,
                "igst": 0,
                "cgst": 9,
                "sgst": 9,
                "cess": 0,
                "document_type": "cdnr",
                "flags": [],
                "source": "GSTR-2B_July.json",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "GSTR_2B_Formatted.xlsx"
            write_gstr_2b(path, rows)
            book = load_workbook(path)
            self.assertIn("CDN", book.sheetnames)
            self.assertEqual(book["CDN"]["C2"].value, "CN-1")
            self.assertIn("Match", _headers(book["CDN"]))


class GstrDumpPackTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        os.environ["LOCALAPPDATA"] = self._tmp.name
        from apps.engine.db import reset_engine

        reset_engine()

    def tearDown(self) -> None:
        from apps.engine.db import reset_engine

        reset_engine()
        self._tmp.cleanup()

    def test_dump_2b_json_writes_b2b_match_empty_acme(self) -> None:
        from apps.engine.clients import create_client
        from apps.engine.dump import ingest_paths, start_job
        from apps.engine.firm import save_firm
        from apps.engine.periods import create_period
        from apps.engine.pipeline import get_period_pack

        save_firm("Test firm")
        client = create_client("Acme")
        period = create_period(client["id"], "Jul 2026")
        job = start_job(period["id"])
        ingest_paths(job["id"], [str(GSTR_2B_JSON)])
        pack = get_period_pack(period["id"])
        self.assertIsNotNone(pack)
        item = next(out for out in pack["outputs"] if out["key"] == "gstr_2b")
        path = Path(item["path"])
        self.assertEqual(path.name, "GSTR_2B_Formatted.xlsx")
        book = load_workbook(path)
        self.assertIn("B2B", book.sheetnames)
        sheet = book["B2B"]
        headers = _headers(sheet)
        self.assertIn("Match", headers)
        invoices = [sheet.cell(row, _col(sheet, "Invoice no")).value for row in range(2, sheet.max_row + 1)]
        self.assertTrue(any(value and "ACME" in str(value) for value in invoices), invoices)
        match_col = _col(sheet, "Match")
        for excel_row in range(2, sheet.max_row + 1):
            self.assertIn(sheet.cell(excel_row, match_col).value, (None, ""))


if __name__ == "__main__":
    unittest.main()
