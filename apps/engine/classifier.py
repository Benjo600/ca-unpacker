from __future__ import annotations

import csv
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

from apps.engine.kinds import IMAGE_SUFFIXES, SPREADSHEET_SUFFIXES, TALLY_SUFFIXES

GSTIN_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][A-Z0-9]Z[A-Z0-9]\b", re.I)

BANK_HINTS = (
    "opening balance",
    "closing balance",
    "withdrawal",
    "deposit",
    "narration",
    "cheque",
    "chq",
    "transaction",
    "debit",
    "credit",
    "account number",
    "ifsc",
)

INVOICE_HINTS = (
    "tax invoice",
    "invoice no",
    "invoice number",
    "irn",
    "hsn",
    "taxable",
    "place of supply",
    "bill of supply",
)

TALLY_MARKERS = ("tallymessage", "<voucher", "daybook", "exportedby", "tallyprime")

ZOHO_HEADERS = {
    "invoice number",
    "gst treatment",
    "place of supply",
    "gst identification number (gstin)",
    "item tax %",
}


@dataclass(frozen=True)
class Classification:
    kind: str
    confidence: float
    reason: str


def classify_path(path: Path) -> Classification:
    name = path.name.lower()
    suffix = path.suffix.lower()

    if suffix == ".json":
        return _classify_json(path, name)
    if suffix == ".zip":
        return _classify_zip(path, name)
    if suffix in TALLY_SUFFIXES:
        return _classify_tally_text(path, name)
    if suffix == ".xls":
        return Classification(
            "unknown",
            0.9,
            "legacy .xls is not supported — export .xlsx or .csv",
        )
    if suffix in SPREADSHEET_SUFFIXES:
        return _classify_spreadsheet(path, name)
    if suffix in IMAGE_SUFFIXES:
        return _classify_image(path, name)
    if suffix == ".pdf":
        return _classify_pdf(path, name)
    return Classification("unknown", 0.2, "no matching file type")


def classify_from_text(name: str, page_count: int, text: str) -> Classification:
    lowered = text.lower()
    bank_hits = sum(1 for hint in BANK_HINTS if hint in lowered)
    invoice_hits = sum(1 for hint in INVOICE_HINTS if hint in lowered)
    has_gstin = bool(GSTIN_RE.search(text.upper()))
    filename = name.lower()

    if _filename_says_gstr(filename, "2b"):
        return Classification("gstr_2b", 0.8, "filename looks like GSTR-2B")
    if _filename_says_gstr(filename, "3b"):
        return Classification("gstr_3b", 0.8, "filename looks like GSTR-3B")
    if _filename_says_gstr(filename, "1") or "gstr-1" in filename or "gstr1" in filename:
        return Classification("gstr_1", 0.75, "filename looks like GSTR-1")

    if bank_hits >= 3 or (bank_hits >= 2 and page_count >= 3):
        return Classification("bank", 0.9, f"bank wording ({bank_hits} hits), {page_count} pages")
    if invoice_hits >= 2 or (has_gstin and ("invoice" in lowered or "tax invoice" in lowered)):
        return Classification("invoice", 0.85, "invoice wording or GSTIN on a short document")
    if page_count >= 3 and bank_hits >= 1:
        return Classification("bank", 0.6, "multi-page PDF with bank wording")
    if page_count <= 2 and (has_gstin or "invoice" in filename):
        return Classification("invoice", 0.55, "short PDF with GSTIN or invoice in the name")
    return Classification("unknown", 0.3, "PDF did not match bank or invoice wording")


def _filename_says_gstr(name: str, token: str) -> bool:
    compact = name.replace(" ", "").replace("_", "").replace("-", "")
    return f"gstr{token}" in compact or f"gst{token}" in compact


def _classify_json(path: Path, name: str) -> Classification:
    if _filename_says_gstr(name, "2b"):
        return Classification("gstr_2b", 0.95, "filename is GSTR-2B JSON")
    if _filename_says_gstr(name, "3b"):
        return Classification("gstr_3b", 0.95, "filename is GSTR-3B JSON")
    if _filename_says_gstr(name, "1") or "gstr-1" in name or "gstr1" in name:
        return Classification("gstr_1", 0.9, "filename is GSTR-1 JSON")

    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        return Classification("unknown", 0.2, "JSON could not be read")

    if not isinstance(payload, dict):
        return Classification("unknown", 0.2, "JSON is not an object")

    data = payload.get("data") if isinstance(payload.get("data"), dict) else payload
    if any(key in data for key in ("docdata", "itcsum", "impsum")) or "docdata" in payload:
        return Classification("gstr_2b", 0.92, "JSON schema looks like GSTR-2B")
    if any(key in payload for key in ("sup_details", "inter_sup", "inward_sup")):
        return Classification("gstr_3b", 0.92, "JSON schema looks like GSTR-3B")
    if any(key in payload for key in ("b2b", "b2cl", "b2cs", "cdnr", "hsn")):
        return Classification("gstr_1", 0.9, "JSON schema looks like GSTR-1")
    return Classification("unknown", 0.35, "JSON did not match a GSTR schema")


def _classify_tally_text(path: Path, name: str) -> Classification:
    if "zoho" in name:
        return Classification("zoho", 0.7, "filename mentions Zoho")
    try:
        sample = path.read_text(encoding="utf-8", errors="ignore")[:8000].lower()
    except OSError:
        return Classification("unknown", 0.2, "file could not be read")
    if any(marker in sample for marker in TALLY_MARKERS):
        return Classification("tally", 0.9, "Tally XML / export markers")
    if "zoho" in sample:
        return Classification("zoho", 0.6, "Zoho mentioned in file")
    return Classification("unknown", 0.3, "text/xml did not look like Tally")


def _classify_zip(path: Path, name: str) -> Classification:
    if "tally" in name or "daybook" in name or "backup" in name:
        guessed = "tally"
    else:
        guessed = None
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            for inner in names[:30]:
                lower = inner.lower()
                if not (lower.endswith(".xml") or lower.endswith(".txt")):
                    continue
                with archive.open(inner) as handle:
                    sample = handle.read(6000).decode("utf-8", errors="ignore").lower()
                if any(marker in sample for marker in TALLY_MARKERS):
                    return Classification("tally", 0.92, f"zip contains Tally export ({Path(inner).name})")
    except (OSError, zipfile.BadZipFile):
        return Classification("unknown", 0.2, "zip could not be read")
    if guessed:
        return Classification("tally", 0.55, "zip name looks like a Tally backup")
    return Classification("unknown", 0.3, "zip did not contain a Tally export")


def _classify_spreadsheet(path: Path, name: str) -> Classification:
    if "zoho" in name or "books" in name:
        return Classification("zoho", 0.8, "filename looks like a Zoho export")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        cells = _xlsx_header_cells(path)
        if cells is None:
            return Classification("unknown", 0.2, "spreadsheet could not be read")
        if _zoho_header_match(cells):
            return Classification("zoho", 0.88, "Excel headers look like Zoho Books")
        return Classification("unknown", 0.35, "spreadsheet is not a Zoho export")
    if suffix != ".csv":
        return Classification("unknown", 0.35, "spreadsheet is not a Zoho CSV")
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader, [])
    except (OSError, UnicodeError):
        return Classification("unknown", 0.2, "CSV could not be read")
    cells = {str(cell).strip().lower() for cell in header if str(cell).strip()}
    if _zoho_header_match(cells):
        return Classification("zoho", 0.88, "CSV headers look like Zoho Books")
    return Classification("unknown", 0.3, "CSV headers did not match Zoho")


def _zoho_header_match(cells: set[str]) -> bool:
    return bool(cells & ZOHO_HEADERS) or "invoice number" in cells


def _xlsx_header_cells(path: Path) -> set[str] | None:
    try:
        from openpyxl import load_workbook

        book = load_workbook(path, read_only=True, data_only=True)
        try:
            first = next(book.active.iter_rows(min_row=1, max_row=1, values_only=True), None)
        finally:
            book.close()
    except Exception:
        return None
    if not first:
        return set()
    return {str(cell).strip().lower() for cell in first if cell is not None and str(cell).strip()}


def _filename_invoice_hint(name: str) -> bool:
    return any(token in name for token in ("inv", "invoice", "bill", "tax"))


def _has_printed_invoice_text(text: str) -> bool:
    if not text or not text.strip():
        return False
    if GSTIN_RE.search(text.upper()):
        return True
    lowered = text.lower()
    if any(hint in lowered for hint in INVOICE_HINTS):
        return True
    return "invoice" in lowered


def _image_too_small(path: Path) -> bool:
    try:
        from PIL import Image

        with Image.open(path) as image:
            width, height = image.size
    except Exception:
        return False
    return width < 8 or height < 8


def _ocr_available() -> bool:
    from apps.engine.ocr import find_tesseract

    if find_tesseract() is None:
        return False
    try:
        import pytesseract  # noqa: F401
    except ImportError:
        return False
    return True


def _ocr_image_text(path: Path) -> str:
    from apps.engine.ocr import ocr_image_lines

    try:
        from PIL import Image

        with Image.open(path) as image:
            width, height = image.size
    except Exception:
        width, height = 1, 1
    if width <= 0 or height <= 0:
        return ""
    lines = ocr_image_lines(path, 1, float(width), float(height), 1.0)
    return "\n".join(line.text for line in lines if line.text)


def _classify_image(path: Path, name: str) -> Classification:
    # 1x1 / tiny bitmaps cannot hold printed GST invoice text.
    if _image_too_small(path):
        return Classification("unknown", 0.4, "image; no printed invoice text")

    if _ocr_available():
        text = _ocr_image_text(path)
        result = classify_from_text(name, 1, text)
        if _has_printed_invoice_text(text):
            if result.kind == "invoice":
                return result
            return Classification("invoice", 0.85, "invoice wording or GSTIN on a short document")
        # OCR ran: filename must not promote junk/handwritten photos to invoice.
        if result.kind == "invoice" or not text.strip():
            return Classification("unknown", 0.4, "image; no printed invoice text")
        return result

    if _filename_invoice_hint(name):
        return Classification("invoice", 0.6, "image name looks like an invoice")
    return Classification("unknown", 0.4, "image; cannot read GSTIN without OCR yet")


def _classify_pdf(path: Path, name: str) -> Classification:
    from apps.engine.pdf_extract import extract_pdf

    extracted = extract_pdf(path)
    if extracted.pdf_type == "encrypted":
        return Classification("unknown", 0.5, "password-protected PDF")
    if extracted.pdf_type in {"scanned", "image_based"}:
        return classify_from_text(name, extracted.page_count, name)
    text = "\n".join(page.text for page in extracted.pages[:4])
    return classify_from_text(name, extracted.page_count, text)
