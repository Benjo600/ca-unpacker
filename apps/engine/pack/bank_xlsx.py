from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Date",
    "Description",
    "Cheque / Ref",
    "Debit",
    "Credit",
    "Balance",
    "Source",
    "Flags",
]

COVER_HEADERS = [
    "File",
    "Bank",
    "Sheet",
    "Account no",
    "IFSC",
    "Rows",
    "Opening",
    "Stated close",
    "Computed close",
    "Result",
    "Broken at",
    "Flags",
]

MONEY_FORMAT = "#,##0.00"
MONEY_INFERRED_FORMAT = '#,##0.00" (inferred)"'
DATE_FORMAT = "DD-MM-YYYY"
RED_FILL = PatternFill("solid", fgColor="FFCDD2")
GREEN_FILL = PatternFill("solid", fgColor="C8E6C9")


def write_bank_pack(path: Path, rows: list[dict], check: dict, meta: dict) -> Path:
    return write_bank_workbook(
        path,
        [
            {
                "rows": rows,
                "check": check,
                "meta": meta,
            }
        ],
    )


def write_bank_workbook(path: Path, files: list[dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    cover = book.active
    cover.title = "Balance Check"

    used: set[str] = {"Balance Check"}
    names: list[str] = []
    if len(files) == 1:
        names.append("Transactions")
        used.add("Transactions")
    else:
        for item in files:
            title = _sheet_name(item["meta"].get("filename") or "Transactions", used)
            used.add(title)
            names.append(title)

    _cover_multi(cover, files, names)

    for item, title in zip(files, names):
        sheet = book.create_sheet(title)
        _write_rows(sheet, item["rows"])

    book.save(path)
    return path


def _write_rows(sheet, rows: list[dict]) -> None:
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        flags = _flag_text(row.get("flags") if row.get("flags") is not None else row.get("validation_flags"))
        sheet.append(
            [
                _as_date(row.get("date")),
                row.get("description"),
                row.get("cheque_ref"),
                _money(row.get("debit")),
                _money(row.get("credit")),
                _money(row.get("balance")),
                row.get("source"),
                flags,
            ]
        )
        excel_row = sheet.max_row
        sheet.cell(excel_row, 1).number_format = DATE_FORMAT
        for col in (4, 5, 6):
            sheet.cell(excel_row, col).number_format = MONEY_FORMAT
        if "balance_mismatch" in _flag_tokens(flags) or "running_balance_break" in _flag_tokens(flags):
            for col in range(1, len(HEADERS) + 1):
                sheet.cell(excel_row, col).fill = RED_FILL
    last_col = get_column_letter(len(HEADERS))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    sheet.freeze_panes = "A2"
    widths = [14, 48, 16, 14, 14, 14, 36, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _sheet_name(filename: str, used: set[str]) -> str:
    stem = Path(filename).stem
    cleaned = re.sub(r"[\[\]\*\?:/\\]", " ", stem)[:28].strip() or "Transactions"
    name = cleaned
    n = 2
    while name in used:
        name = f"{cleaned[:26]}_{n}"
        n += 1
    return name


def _cover_multi(sheet, files: list[dict], names: list[str]) -> None:
    sheet["A1"] = "Bank statement pack"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = "Each source file is checked on its own. Do not mix two statements when you test."
    sheet.append([])
    sheet.append(COVER_HEADERS)
    for cell in sheet[4]:
        cell.font = Font(bold=True)
    for item, title in zip(files, names):
        check = item["check"]
        meta = item["meta"]
        inferred = bool(check.get("opening_inferred"))
        status = str(check.get("status") or "")
        if status == "match":
            label = "MATCH"
            fill = GREEN_FILL
        elif status == "unverified":
            label = "COULD NOT VERIFY"
            fill = None
        else:
            label = "MISMATCH"
            fill = RED_FILL
        sheet.append(
            [
                meta.get("filename"),
                meta.get("profile_label"),
                title,
                meta.get("account_number"),
                meta.get("ifsc"),
                check.get("row_count"),
                check.get("opening_balance"),
                check.get("stated_closing"),
                check.get("computed_closing"),
                label,
                check.get("broken_at_row"),
                _flag_text(check.get("flags")),
            ]
        )
        excel_row = sheet.max_row
        if fill is not None:
            for col in range(1, len(COVER_HEADERS) + 1):
                sheet.cell(excel_row, col).fill = fill
        sheet.cell(excel_row, 7).number_format = MONEY_INFERRED_FORMAT if inferred else MONEY_FORMAT
        sheet.cell(excel_row, 8).number_format = MONEY_FORMAT
        sheet.cell(excel_row, 9).number_format = MONEY_FORMAT
    if any(item["check"].get("opening_inferred") for item in files):
        sheet.cell(4, 7).value = "Opening (inferred)"
    widths = [36, 22, 18, 18, 14, 10, 18, 16, 16, 12, 12, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _as_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return value


def _money(value):
    if value is None or value == "":
        return None
    return float(value)


def _flag_text(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return value
    return ", ".join(str(item) for item in value if item)


def _flag_tokens(value) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, str):
        return [part.strip() for part in value.split(",") if part.strip()]
    return [str(item) for item in value if item]
