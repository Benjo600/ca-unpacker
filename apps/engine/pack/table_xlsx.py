from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY_FORMAT = "#,##0.00"
MONEY_KEYS = {"taxable", "tax", "invoice_value", "taxable_value", "amount", "rate"}
FLAG_MARKERS = ("gstin_checksum", "hsn_length", "invoice_math")
RED_FILL = PatternFill("solid", fgColor="FFCDD2")


def write_table(path: Path, sheet_name: str, rows: list[dict], columns: list[tuple[str, str]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    _fill_sheet(sheet, sheet_name, rows, columns)
    book.save(path)
    return path


def write_purchase_workbook(
    path: Path,
    invoices: list[dict],
    line_items: list[dict],
    invoice_cols: list[tuple[str, str]],
    line_cols: list[tuple[str, str]],
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    invoices_sheet = book.active
    _fill_sheet(invoices_sheet, "Invoices", invoices, invoice_cols)
    items_sheet = book.create_sheet("Line items")
    _fill_sheet(items_sheet, "Line items", line_items, line_cols)
    book.save(path)
    return path


def _fill_sheet(sheet, sheet_name: str, rows: list[dict], columns: list[tuple[str, str]]) -> None:
    sheet.title = sheet_name[:31]
    sheet.append([title for title, _key in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    flag_col = next(
        (index for index, (title, key) in enumerate(columns, start=1) if key == "flags" or title == "Flags"),
        None,
    )
    for row in rows:
        sheet.append([_cell(row.get(key)) for _title, key in columns])
        excel_row = sheet.max_row
        for index, (_title, key) in enumerate(columns, start=1):
            if key in MONEY_KEYS:
                sheet.cell(excel_row, index).number_format = MONEY_FORMAT
        if flag_col is not None:
            flag_text = str(sheet.cell(excel_row, flag_col).value or "")
            if any(marker in flag_text for marker in FLAG_MARKERS):
                for col in range(1, len(columns) + 1):
                    sheet.cell(excel_row, col).fill = RED_FILL
    for index, (_title, key) in enumerate(columns, start=1):
        width = 16
        if key in {"description", "trade_name", "party_name", "supplier_name", "raw_excerpt"}:
            width = 40
        if key == "source":
            width = 28
        sheet.column_dimensions[get_column_letter(index)].width = width
    if not rows:
        sheet.append(["No rows extracted"])
    last_col = get_column_letter(max(len(columns), 1))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    sheet.freeze_panes = "A2"


def _cell(value):
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return value
