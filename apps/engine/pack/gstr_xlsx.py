from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY_FORMAT = "#,##0.00"
RED_FILL = PatternFill("solid", fgColor="FFCDD2")

B2B_HEADERS = [
    "GSTIN",
    "Trade name",
    "Invoice no",
    "Date",
    "Value",
    "Taxable",
    "IGST",
    "CGST",
    "SGST",
    "Cess",
    "ITC",
    "Type",
    "Flags",
    "Source",
    "Match",
    "Books ref",
]
B2B_MONEY_COLS = (5, 6, 7, 8, 9, 10)  # Value through Cess

HSN_HEADERS = ["HSN", "Taxable", "IGST", "CGST", "SGST", "Cess", "Flags", "Source"]
HSN_MONEY_COLS = (2, 3, 4, 5, 6)

SUMMARY_HEADERS = ["Section", "Taxable", "IGST", "CGST", "SGST", "Cess", "Source"]
SUMMARY_MONEY_COLS = (2, 3, 4, 5, 6)

CDN_TYPES = {"CDN", "CDNR", "CDNA", "CDNRA"}
MATCH_NOTE = "Match columns are empty until reconciliation."
MATCH_FILLED_NOTE = "Match columns filled from the master reconciliation grid."


def write_gstr_2b(path: Path, rows: list[dict], meta: dict | None = None) -> Path:
    b2b = [row for row in rows if not _is_cdn(row)]
    cdn = [row for row in rows if _is_cdn(row)]
    flagged = [row for row in rows if _is_flagged(row)]
    book = _new_book("Cover")
    _write_cover(
        book.active,
        "GSTR-2B",
        rows,
        meta,
        [
            ("B2B rows", len(b2b)),
            ("CDN rows", len(cdn)),
            ("Flagged rows", len(flagged)),
            ("Total rows", len(rows)),
        ],
        note=_match_note(rows),
    )
    _write_b2b_sheet(book.create_sheet("B2B"), b2b, default_type="B2B")
    if cdn:
        _write_b2b_sheet(book.create_sheet("CDN"), cdn, default_type="CDN")
    _write_flags_sheet(book.create_sheet("Flags"), flagged)
    return _save(book, path)


def write_gstr_1(path: Path, rows: list[dict], meta: dict | None = None) -> Path:
    b2b = [row for row in rows if not _is_hsn(row)]
    hsn = [row for row in rows if _is_hsn(row)]
    flagged = [row for row in rows if _is_flagged(row)]
    book = _new_book("Cover")
    counts = [
        ("B2B rows", len(b2b)),
        ("HSN rows", len(hsn)),
        ("Flagged rows", len(flagged)),
        ("Total rows", len(rows)),
    ]
    _write_cover(book.active, "GSTR-1", rows, meta, counts, note=MATCH_NOTE)
    _write_b2b_sheet(book.create_sheet("B2B"), b2b, default_type="B2B")
    if hsn:
        _write_hsn_sheet(book.create_sheet("HSN"), hsn)
    _write_flags_sheet(book.create_sheet("Flags"), flagged)
    return _save(book, path)


def write_gstr_3b(path: Path, rows: list[dict], meta: dict | None = None) -> Path:
    book = _new_book("Cover")
    _write_cover(
        book.active,
        "GSTR-3B",
        rows,
        meta,
        [("Summary rows", len(rows))],
        note=None,
    )
    _write_summary_sheet(book.create_sheet("Summary"), rows)
    return _save(book, path)


def _new_book(title: str) -> Workbook:
    book = Workbook()
    book.active.title = title[:31]
    return book


def _save(book: Workbook, path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def _write_cover(sheet, title: str, rows: list[dict], meta: dict | None, counts: list[tuple[str, int]], note: str | None) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["GSTIN", _cover_gstin(rows, meta)])
    sheet.append(["Period", _cover_period(rows, meta)])
    sheet.append(["Source", _cover_source(rows)])
    for label, value in counts:
        sheet.append([label, value])
    if note:
        sheet.append([])
        sheet.append([note])
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 28
    sheet.freeze_panes = "A2"


def _write_b2b_sheet(sheet, rows: list[dict], default_type: str) -> None:
    _write_header(sheet, B2B_HEADERS)
    for row in rows:
        values = _b2b_values(row, _row_type(row, default_type))
        sheet.append(values)
        excel_row = sheet.max_row
        for col in B2B_MONEY_COLS:
            sheet.cell(excel_row, col).number_format = MONEY_FORMAT
        if _is_flagged(row):
            _paint_row(sheet, excel_row, len(B2B_HEADERS))
    if not rows:
        sheet.append(["No rows extracted"])
    _finish_table(sheet, len(B2B_HEADERS), [14, 28, 20, 14, 14, 14, 12, 12, 12, 12, 12, 10, 22, 28, 14, 16])


def _write_hsn_sheet(sheet, rows: list[dict]) -> None:
    _write_header(sheet, HSN_HEADERS)
    for row in rows:
        sheet.append(
            [
                row.get("hsn") or row.get("invoice_number"),
                _money(row.get("taxable")),
                _money(row.get("igst")),
                _money(row.get("cgst")),
                _money(row.get("sgst")),
                _money(row.get("cess")),
                _flag_text(row.get("flags")),
                row.get("source"),
            ]
        )
        excel_row = sheet.max_row
        for col in HSN_MONEY_COLS:
            sheet.cell(excel_row, col).number_format = MONEY_FORMAT
    _finish_table(sheet, len(HSN_HEADERS), [14, 14, 12, 12, 12, 12, 22, 28])


def _write_summary_sheet(sheet, rows: list[dict]) -> None:
    _write_header(sheet, SUMMARY_HEADERS)
    for row in rows:
        sheet.append(
            [
                row.get("section"),
                _money(row.get("taxable")),
                _money(row.get("igst")),
                _money(row.get("cgst")),
                _money(row.get("sgst")),
                _money(row.get("cess")),
                row.get("source"),
            ]
        )
        excel_row = sheet.max_row
        for col in SUMMARY_MONEY_COLS:
            sheet.cell(excel_row, col).number_format = MONEY_FORMAT
    if not rows:
        sheet.append(["No rows extracted"])
    _finish_table(sheet, len(SUMMARY_HEADERS), [44, 14, 12, 12, 12, 12, 28])


def _write_flags_sheet(sheet, rows: list[dict]) -> None:
    _write_header(sheet, B2B_HEADERS)
    if not rows:
        sheet.append(["none"])
    else:
        for row in rows:
            sheet.append(_b2b_values(row, _row_type(row, "B2B")))
            excel_row = sheet.max_row
            for col in B2B_MONEY_COLS:
                sheet.cell(excel_row, col).number_format = MONEY_FORMAT
            _paint_row(sheet, excel_row, len(B2B_HEADERS))
    _finish_table(sheet, len(B2B_HEADERS), [14, 28, 20, 14, 14, 14, 12, 12, 12, 12, 12, 10, 22, 28, 14, 16])


def _match_note(rows: list[dict]) -> str:
    for row in rows:
        if str(row.get("match_status") or "").strip():
            return MATCH_FILLED_NOTE
    return MATCH_NOTE


def _write_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _finish_table(sheet, ncols: int, widths: list[int]) -> None:
    last_col = get_column_letter(max(ncols, 1))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    sheet.freeze_panes = "A2"
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _paint_row(sheet, excel_row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        sheet.cell(excel_row, col).fill = RED_FILL


def _b2b_values(row: dict, doc_type: str) -> list:
    return [
        row.get("gstin"),
        row.get("trade_name"),
        row.get("invoice_number"),
        row.get("invoice_date"),
        _money(row.get("invoice_value")),
        _money(row.get("taxable")),
        _money(row.get("igst")),
        _money(row.get("cgst")),
        _money(row.get("sgst")),
        _money(row.get("cess")),
        row.get("itc_availability") if row.get("itc_availability") is not None else row.get("itc"),
        doc_type,
        _flag_text(row.get("flags")),
        row.get("source"),
        row.get("match_status") or None,
        row.get("books_ref") or None,
    ]


def _cover_gstin(rows: list[dict], meta: dict | None) -> str | None:
    if meta:
        for key in ("gstin", "filer_gstin"):
            if meta.get(key):
                return meta[key]
    for row in rows:
        for key in ("filer_gstin", "recipient_gstin"):
            if row.get(key):
                return row[key]
    return None


def _cover_period(rows: list[dict], meta: dict | None) -> str | None:
    if meta:
        for key in ("period", "rtnprd", "fp", "ret_period"):
            if meta.get(key):
                return meta[key]
    for row in rows:
        for key in ("period", "rtnprd", "fp", "ret_period"):
            if row.get(key):
                return row[key]
    return None


def _cover_source(rows: list[dict]) -> str | None:
    if not rows:
        return None
    source = str(rows[0].get("source") or "")
    return source.split("#", 1)[0] or None


def _row_type(row: dict, default: str) -> str:
    raw = row.get("document_type")
    if raw is None or str(raw).strip() == "":
        if _looks_hsn(row):
            return "HSN"
        return default
    return str(raw).strip()


def _is_hsn(row: dict) -> bool:
    return _row_type(row, "B2B").upper() == "HSN" or _looks_hsn(row)


def _looks_hsn(row: dict) -> bool:
    source = str(row.get("source") or "")
    if "#hsn" in source.lower():
        return True
    if str(row.get("trade_name") or "").strip().lower() == "hsn summary":
        return True
    return str(row.get("document_type") or "").strip().upper() == "HSN"


def _is_cdn(row: dict) -> bool:
    return str(row.get("document_type") or "").strip().upper() in CDN_TYPES


def _is_flagged(row: dict) -> bool:
    flags = row.get("flags")
    if flags is None or flags == "":
        return False
    if isinstance(flags, str):
        return bool(flags.strip())
    return any(str(item).strip() for item in flags)


def _flag_text(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return value
    return ", ".join(str(item) for item in value if item)


def _money(value):
    if value is None or value == "":
        return None
    return float(value)
