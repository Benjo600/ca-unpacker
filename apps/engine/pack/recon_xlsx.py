from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY_FORMAT = "#,##0.00"
RED_FILL = PatternFill("solid", fgColor="FFCDD2")
BANK_NOTE = "Bank hints are an assist only; they are not matches."
COUNT_ORDER = (
    ("matched", "matched"),
    ("books_only", "books_only"),
    ("portal_only", "portal_only"),
    ("amount_mismatch", "amount_mismatch"),
    ("likely", "likely"),
)
GRID_HEADERS = [
    "Status",
    "GSTIN",
    "Party",
    "Invoice no (2B)",
    "Invoice no (books)",
    "Date (2B)",
    "Date (books)",
    "Amount (2B)",
    "Amount (books)",
    "Amount diff",
    "Bank hint",
    "Source (2B)",
    "Source (books)",
]
GRID_KEYS = [
    "status",
    "gstin",
    "party",
    "invoice_2b",
    "invoice_books",
    "date_2b",
    "date_books",
    "amount_2b",
    "amount_books",
    "amount_diff",
    "bank_hint",
    "source_2b",
    "source_books",
]
MONEY_COLS = (8, 9, 10)
UNMATCHED_STATUSES = {"books_only", "portal_only", "amount_mismatch", "likely"}
FLAG_STATUSES = {"books_only", "portal_only", "amount_mismatch", "likely"}
WIDTHS = [16, 18, 28, 20, 20, 14, 14, 14, 16, 14, 28, 28, 28]


def write_master_grid(path: Path, result: dict) -> Path:
    rows = list((result or {}).get("rows") or [])
    counts = (result or {}).get("counts") or {}
    book = Workbook()
    cover = book.active
    cover.title = "Cover"
    _write_cover(cover, counts, len(rows))
    _write_table(book.create_sheet("Grid"), rows)
    unmatched = [row for row in rows if str(row.get("status") or "") in UNMATCHED_STATUSES]
    _write_table(book.create_sheet("Unmatched"), unmatched)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def _write_cover(sheet, counts: dict, row_count: int) -> None:
    sheet["A1"] = "Master reconciliation"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    for key, label in COUNT_ORDER:
        sheet.append([label, int(counts.get(key) or 0)])
    sheet.append(["Total rows", row_count])
    sheet.append([])
    sheet.append([BANK_NOTE])
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 16
    sheet.freeze_panes = "A2"


def _write_table(sheet, rows: list[dict]) -> None:
    sheet.append(GRID_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        values = [row.get(key) for key in GRID_KEYS]
        sheet.append(values)
        excel_row = sheet.max_row
        for col in MONEY_COLS:
            sheet.cell(excel_row, col).number_format = MONEY_FORMAT
        status = str(row.get("status") or "")
        if status in FLAG_STATUSES:
            for col in range(1, len(GRID_HEADERS) + 1):
                sheet.cell(excel_row, col).fill = RED_FILL
    if not rows:
        sheet.append(["No rows"])
    last_col = get_column_letter(len(GRID_HEADERS))
    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"
    sheet.freeze_panes = "A2"
    for index, width in enumerate(WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
